from fastapi import FastAPI, APIRouter, WebSocket, WebSocketDisconnect, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import asyncio
import json
from pyais import decode
from pyais.stream import TCPConnection, UDPReceiver, FileReaderStream
import socket
import serial
import serial.tools.list_ports
import threading
from bson import ObjectId
from openpyxl import Workbook
from io import BytesIO
import math

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# WebSocket connections manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()

# Background stream controllers
active_streams = {}
main_event_loop = None

# Models
class VesselPosition(BaseModel):
    mmsi: str
    timestamp: datetime
    lat: float
    lon: float
    speed: Optional[float] = None
    course: Optional[float] = None
    heading: Optional[int] = None
    nav_status: Optional[int] = None

class VesselInfo(BaseModel):
    mmsi: str
    name: Optional[str] = None
    callsign: Optional[str] = None
    imo: Optional[int] = None
    ship_type: Optional[int] = None
    ship_type_text: Optional[str] = None
    dimension_a: Optional[int] = None
    dimension_b: Optional[int] = None
    dimension_c: Optional[int] = None
    dimension_d: Optional[int] = None
    last_seen: Optional[datetime] = None
    destination: Optional[str] = None
    eta: Optional[str] = None

class AISMessage(BaseModel):
    mmsi: str
    timestamp: datetime
    message_type: int
    raw: str
    decoded: Dict[str, Any]

class StreamConfig(BaseModel):
    stream_type: str  # tcp, udp, serial
    host: Optional[str] = None
    port: Optional[int] = None
    serial_port: Optional[str] = None
    baudrate: Optional[int] = 9600

class SearchQuery(BaseModel):
    mmsi: Optional[str] = None
    vessel_name: Optional[str] = None
    ship_type: Optional[int] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None

class DataSource(BaseModel):
    source_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source_type: str  # tcp, udp, serial, file
    name: str
    config: Dict[str, Any]
    status: str = "active"  # active, inactive, paused
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_message: Optional[datetime] = None
    message_count: int = 0
    target_count: int = 0  # Number of unique vessels
    fragment_count: int = 0  # Number of omitted/incomplete messages
    message_limit: int = 500  # Max messages to keep per source
    target_limit: int = 0  # Max targets to display (0 = unlimited)
    spoof_limit_km: float = 500.0  # Default 500km spoof limit
    is_paused: bool = False  # Pause state for streaming sources

# Helper functions
def serialize_doc(doc):
    """Convert MongoDB document to JSON-serializable dict"""
    if doc is None:
        return None
    
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    
    if isinstance(doc, dict):
        serialized = {}
        for key, value in doc.items():
            if isinstance(value, ObjectId):
                serialized[key] = str(value)
            elif isinstance(value, dict):
                serialized[key] = serialize_doc(value)
            elif isinstance(value, list):
                serialized[key] = [serialize_doc(item) if isinstance(item, dict) else item for item in value]
            else:
                serialized[key] = value
        return serialized
    
    return doc

def get_mmsi_country(mmsi: str) -> str:
    """Get country from MMSI based on MID (Maritime Identification Digits)"""
    if not mmsi or len(mmsi) < 3:
        return "Unknown"
    
    # Extract MID (first 3 digits)
    mid = mmsi[:3]
    
    # Comprehensive MID to country mapping
    mid_countries = {
        "201": "Albania", "202": "Andorra", "203": "Austria", "204": "Azores",
        "205": "Belgium", "206": "Belarus", "207": "Bulgaria", "208": "Vatican",
        "209": "Cyprus", "210": "Cyprus", "211": "Germany", "212": "Cyprus",
        "213": "Georgia", "214": "Moldova", "215": "Malta", "216": "Armenia",
        "218": "Germany", "219": "Denmark", "220": "Denmark", "224": "Spain",
        "225": "Spain", "226": "France", "227": "France", "228": "France",
        "229": "Malta", "230": "Finland", "231": "Faroe Islands", "232": "United Kingdom",
        "233": "United Kingdom", "234": "United Kingdom", "235": "United Kingdom",
        "236": "Gibraltar", "237": "Greece", "238": "Croatia", "239": "Greece",
        "240": "Greece", "241": "Greece", "242": "Morocco", "243": "Hungary",
        "244": "Netherlands", "245": "Netherlands", "246": "Netherlands", "247": "Italy",
        "248": "Malta", "249": "Malta", "250": "Ireland", "251": "Iceland",
        "252": "Liechtenstein", "253": "Luxembourg", "254": "Monaco", "255": "Madeira",
        "256": "Malta", "257": "Norway", "258": "Norway", "259": "Norway",
        "261": "Poland", "262": "Montenegro", "263": "Portugal", "264": "Romania",
        "265": "Sweden", "266": "Sweden", "267": "Slovak Republic", "268": "San Marino",
        "269": "Switzerland", "270": "Czech Republic", "271": "Turkey", "272": "Ukraine",
        "273": "Russia", "274": "North Macedonia", "275": "Latvia", "276": "Estonia",
        "277": "Lithuania", "278": "Slovenia", "279": "Serbia", "301": "Anguilla",
        "303": "USA", "304": "Antigua and Barbuda", "305": "Antigua and Barbuda",
        "306": "Curaçao", "307": "Aruba", "308": "Bahamas", "309": "Bahamas",
        "310": "Bermuda", "311": "Bahamas", "312": "Belize", "314": "Barbados",
        "316": "Canada", "319": "Cayman Islands", "321": "Costa Rica", "323": "Cuba",
        "325": "Dominica", "327": "Dominican Republic", "329": "Guadeloupe",
        "330": "Grenada", "331": "Greenland", "332": "Guatemala", "334": "Honduras",
        "336": "Haiti", "338": "USA", "339": "Jamaica", "341": "St Kitts and Nevis",
        "343": "St Lucia", "345": "Mexico", "347": "Martinique", "348": "Montserrat",
        "350": "Nicaragua", "351": "Panama", "352": "Panama", "353": "Panama",
        "354": "Panama", "355": "Panama", "356": "Panama", "357": "Panama",
        "358": "Puerto Rico", "359": "El Salvador", "361": "St Pierre and Miquelon",
        "362": "Trinidad and Tobago", "364": "Turks and Caicos Islands",
        "366": "USA", "367": "USA", "368": "USA", "369": "USA",
        "370": "Panama", "371": "Panama", "372": "Panama", "373": "Panama",
        "374": "Panama", "375": "St Vincent and Grenadines", "376": "St Vincent and Grenadines",
        "377": "St Vincent and Grenadines", "378": "British Virgin Islands",
        "379": "US Virgin Islands", "401": "Afghanistan", "403": "Saudi Arabia",
        "405": "Bangladesh", "408": "Bahrain", "410": "Bhutan", "412": "China",
        "413": "China", "414": "China", "416": "Taiwan", "417": "Sri Lanka",
        "419": "India", "422": "Iran", "423": "Azerbaijan", "425": "Iraq",
        "428": "Israel", "431": "Japan", "432": "Japan", "434": "Turkmenistan",
        "436": "Kazakhstan", "437": "Uzbekistan", "438": "Jordan", "440": "South Korea",
        "441": "South Korea", "443": "Palestine", "445": "North Korea", "447": "Kuwait",
        "450": "Lebanon", "451": "Kyrgyzstan", "453": "Macao", "455": "Maldives",
        "457": "Mongolia", "459": "Nepal", "461": "Oman", "463": "Pakistan",
        "466": "Qatar", "468": "Syria", "470": "UAE", "471": "UAE",
        "472": "Tajikistan", "473": "Yemen", "475": "Yemen", "477": "Hong Kong",
        "478": "Bosnia and Herzegovina", "501": "Antarctica", "503": "Australia",
        "506": "Myanmar", "508": "Brunei", "510": "Micronesia", "511": "Palau",
        "512": "New Zealand", "514": "Cambodia", "515": "Cambodia", "516": "Christmas Island",
        "518": "Cook Islands", "520": "Fiji", "523": "Cocos Islands", "525": "Indonesia",
        "529": "Kiribati", "531": "Laos", "533": "Malaysia", "536": "Northern Mariana Islands",
        "538": "Marshall Islands", "540": "New Caledonia", "542": "Niue", "544": "Nauru",
        "546": "French Polynesia", "548": "Philippines", "550": "Timor-Leste",
        "553": "Papua New Guinea", "555": "Pitcairn Island", "557": "Solomon Islands",
        "559": "American Samoa", "561": "Samoa", "563": "Singapore", "564": "Singapore",
        "565": "Singapore", "566": "Singapore", "567": "Thailand", "570": "Tonga",
        "572": "Tuvalu", "574": "Vietnam", "576": "Vanuatu", "577": "Vanuatu",
        "578": "Wallis and Futuna", "601": "South Africa", "603": "Angola",
        "605": "Algeria", "607": "St Paul and Amsterdam Islands", "608": "Ascension Island",
        "609": "Burundi", "610": "Benin", "611": "Botswana", "612": "Central African Republic",
        "613": "Cameroon", "615": "Congo", "616": "Comoros", "617": "Cape Verde",
        "618": "Crozet Archipelago", "619": "Ivory Coast", "620": "Comoros",
        "621": "Djibouti", "622": "Egypt", "624": "Ethiopia", "625": "Eritrea",
        "626": "Gabon", "627": "Ghana", "629": "Gambia", "630": "Guinea-Bissau",
        "631": "Equatorial Guinea", "632": "Guinea", "633": "Burkina Faso",
        "634": "Kenya", "635": "Kerguelen Islands", "636": "Liberia", "637": "Liberia",
        "638": "South Sudan", "642": "Libya", "644": "Lesotho", "645": "Mauritius",
        "647": "Madagascar", "649": "Mali", "650": "Mozambique", "654": "Mauritania",
        "655": "Malawi", "656": "Niger", "657": "Nigeria", "659": "Namibia",
        "660": "Reunion", "661": "Rwanda", "662": "Sudan", "663": "Senegal",
        "664": "Seychelles", "665": "St Helena", "666": "Somalia", "667": "Sierra Leone",
        "668": "Sao Tome and Principe", "669": "Swaziland", "670": "Chad",
        "671": "Togo", "672": "Tunisia", "674": "Tanzania", "675": "Uganda",
        "676": "DR Congo", "677": "Tanzania", "678": "Zambia", "679": "Zimbabwe",
        "701": "Argentina", "710": "Brazil", "720": "Bolivia", "725": "Chile",
        "730": "Colombia", "735": "Ecuador", "740": "Falkland Islands", "745": "Guiana",
        "750": "Guyana", "755": "Paraguay", "760": "Peru", "765": "Suriname",
        "770": "Uruguay", "775": "Venezuela"
    }
    
    return mid_countries.get(mid, f"Unknown ({mid})")

def get_ship_type_text(ship_type: int) -> str:
    ship_types = {
        0: "Not available",
        20: "Wing in ground",
        29: "Wing in ground (Hazardous)",
        30: "Fishing",
        31: "Towing",
        32: "Towing: length exceeds 200m",
        33: "Dredging or underwater ops",
        34: "Diving ops",
        35: "Military ops",
        36: "Sailing",
        37: "Pleasure Craft",
        40: "High speed craft",
        41: "High speed craft (Hazardous)",
        50: "Pilot Vessel",
        51: "Search and Rescue vessel",
        52: "Tug",
        53: "Port Tender",
        54: "Anti-pollution equipment",
        55: "Law Enforcement",
        58: "Medical Transport",
        59: "Special Craft",
        60: "Passenger",
        69: "Passenger (Hazardous)",
        70: "Cargo",
        79: "Cargo (Hazardous)",
        80: "Tanker",
        89: "Tanker (Hazardous)",
        90: "Other Type",
    }
    return ship_types.get(ship_type, f"Unknown ({ship_type})")

def is_valid_position(lat: float, lon: float) -> bool:
    """
    Validate if position is within valid ranges.
    Invalid positions (e.g., 181, 91) are used to indicate no valid position data.
    
    Returns:
        True if position is valid, False otherwise
    """
    if lat is None or lon is None:
        return False
    
    # Valid ranges: lat [-90, 90], lon [-180, 180]
    if lat < -90 or lat > 90:
        return False
    if lon < -180 or lon > 180:
        return False
    
    return True

async def get_last_valid_position(mmsi: str) -> Optional[Dict[str, float]]:
    """
    Get the last valid position for a vessel.
    
    Returns:
        Dictionary with 'lat' and 'lon' keys, or None if no valid position found
    """
    # Find the most recent valid position for this MMSI
    last_valid = await db.positions.find_one(
        {
            'mmsi': mmsi,
            'position_valid': True
        },
        sort=[('timestamp', -1)]
    )
    
    if last_valid:
        return {
            'lat': last_valid.get('display_lat'),
            'lon': last_valid.get('display_lon')
        }
    return None

async def backfill_invalid_positions(mmsi: str, valid_lat: float, valid_lon: float):
    """
    Backfill any previous invalid positions with the first valid position coordinates.
    This ensures smooth trails without position jumps.
    """
    # Find all positions for this MMSI that are invalid and don't have display coordinates
    invalid_positions = await db.positions.find(
        {
            'mmsi': mmsi,
            'position_valid': False,
            'display_lat': {'$exists': False}
        }
    ).to_list(None)
    
    if invalid_positions:
        # Update all invalid positions with the valid coordinates
        for pos in invalid_positions:
            await db.positions.update_one(
                {'_id': pos['_id']},
                {
                    '$set': {
                        'display_lat': valid_lat,
                        'display_lon': valid_lon,
                        'backfilled': True
                    }
                }
            )
        logger.info(f"Backfilled {len(invalid_positions)} invalid positions for MMSI {mmsi}")

async def process_ais_message(raw_message: str, source: str = "unknown", source_id: str = None):
    """Process and store AIS message"""
    try:
        # Check if source is paused
        if source_id:
            source_doc = await db.sources.find_one({'source_id': source_id})
            if source_doc and source_doc.get('is_paused', False):
                logger.debug(f"Source {source_id} is paused, skipping message")
                return
        
        # Decode the message
        decoded_msg = decode(raw_message)
        
        if not decoded_msg:
            # Count as fragment/omitted message
            if source_id:
                await db.sources.update_one(
                    {'source_id': source_id},
                    {'$inc': {'fragment_count': 1}}
                )
            return
        
        # Convert to dict
        decoded = decoded_msg.asdict()
        
        mmsi = str(decoded.get('mmsi', 'unknown'))
        msg_type = decoded.get('msg_type', 0)
        timestamp = datetime.now(timezone.utc)
        
        # Determine if VDO or VDM
        # VDO format: !xxVDO or $xxVDO (where xx is 2-char talker ID like AB, AI, etc.)
        # VDM format: !xxVDM or $xxVDM
        is_vdo = 'VDO' in raw_message[:10]  # Check first 10 chars for VDO
        
        # Store raw message
        message_doc = {
            'mmsi': mmsi,
            'timestamp': timestamp.isoformat(),
            'message_type': msg_type,
            'raw': raw_message,
            'decoded': decoded,
            'source': source,
            'source_id': source_id,
            'is_vdo': is_vdo,
            'repeat_indicator': decoded.get('repeat', 0)
        }
        await db.messages.insert_one(message_doc)
        
        # Purge old messages if limit exceeded
        if source_id:
            source_doc = await db.sources.find_one({'source_id': source_id})
            if source_doc:
                message_limit = source_doc.get('message_limit', 500)
                message_count = await db.messages.count_documents({'source_id': source_id})
                
                if message_count > message_limit:
                    # Delete oldest messages beyond limit
                    messages_to_delete = message_count - message_limit
                    old_messages = await db.messages.find(
                        {'source_id': source_id}
                    ).sort('timestamp', 1).limit(messages_to_delete).to_list(messages_to_delete)
                    
                    if old_messages:
                        message_ids = [msg['_id'] for msg in old_messages]
                        await db.messages.delete_many({'_id': {'$in': message_ids}})
                        
                        # Also delete associated positions
                        position_ids = [msg.get('mmsi') for msg in old_messages]
                        await db.positions.delete_many({
                            'source_id': source_id,
                            'mmsi': {'$in': position_ids},
                            'timestamp': {'$lte': old_messages[-1]['timestamp']}
                        })
                        
                        logger.info(f"Purged {messages_to_delete} old messages from source {source_id}")
        
        # Process based on message type
        if msg_type in [1, 2, 3]:  # Position Reports (Class A)
            original_lat = decoded.get('lat')
            original_lon = decoded.get('lon')
            
            # Validate position
            position_is_valid = is_valid_position(original_lat, original_lon)
            
            # Determine display coordinates
            if position_is_valid:
                display_lat = original_lat
                display_lon = original_lon
            else:
                # Try to get last valid position
                last_valid = await get_last_valid_position(mmsi)
                if last_valid:
                    display_lat = last_valid['lat']
                    display_lon = last_valid['lon']
                else:
                    # No previous valid position, leave as None for now
                    display_lat = None
                    display_lon = None
            
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': original_lat,
                'lon': original_lon,
                'display_lat': display_lat,
                'display_lon': display_lon,
                'position_valid': position_is_valid,
                'speed': decoded.get('speed'),
                'course': decoded.get('course'),
                'heading': decoded.get('heading'),
                'nav_status': decoded.get('status'),
                'source_id': source_id,
                'is_vdo': is_vdo,
                'repeat_indicator': decoded.get('repeat', 0)
            }
            await db.positions.insert_one(position_doc)
            
            # If this is the first valid position, backfill any previous invalid positions
            if position_is_valid:
                await backfill_invalid_positions(mmsi, display_lat, display_lon)
            
            # Get position count
            pos_count = await db.positions.count_documents({'mmsi': mmsi})
            
            # Only update vessel's last_position if we have valid display coordinates
            if display_lat is not None and display_lon is not None:
                # Update vessel last position and add to sources list
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_position': position_doc,
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi)
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
                
                # Broadcast to WebSocket clients
                await manager.broadcast({
                    'type': 'position',
                    'data': position_doc
                })
            else:
                # No valid position to display yet, just update metadata
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi)
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
        
        elif msg_type == 4:  # Base Station Report (VDO)
            # Type 4 has position data but different fields than Type 1-3
            original_lat = decoded.get('lat')
            original_lon = decoded.get('lon')
            
            # Validate position
            position_is_valid = is_valid_position(original_lat, original_lon)
            
            # Determine display coordinates
            if position_is_valid:
                display_lat = original_lat
                display_lon = original_lon
            else:
                # Try to get last valid position
                last_valid = await get_last_valid_position(mmsi)
                if last_valid:
                    display_lat = last_valid['lat']
                    display_lon = last_valid['lon']
                else:
                    # No previous valid position, leave as None for now
                    display_lat = None
                    display_lon = None
            
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': original_lat,
                'lon': original_lon,
                'display_lat': display_lat,
                'display_lon': display_lon,
                'position_valid': position_is_valid,
                'accuracy': decoded.get('accuracy'),
                'source_id': source_id,
                'is_vdo': is_vdo,
                'repeat_indicator': decoded.get('repeat', 0),
                'epfd': decoded.get('epfd'),  # Electronic Position Fixing Device type
                'raim': decoded.get('raim')  # RAIM flag
            }
            
            # Always store the position (even if invalid for data integrity)
            await db.positions.insert_one(position_doc)
            
            # If this is the first valid position, backfill any previous invalid positions
            if position_is_valid:
                await backfill_invalid_positions(mmsi, display_lat, display_lon)
            
            # Get position count
            pos_count = await db.positions.count_documents({'mmsi': mmsi})
            
            # Only update vessel's last_position if we have valid display coordinates
            if display_lat is not None and display_lon is not None:
                # Update vessel last position and add to sources list
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_position': position_doc,
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi),
                            'is_base_station': True  # Mark as base station
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
                
                # Broadcast to WebSocket clients
                await manager.broadcast({
                    'type': 'position',
                    'data': position_doc
                })
            else:
                # No valid position to display yet, just update metadata
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi),
                            'is_base_station': True
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
        
        elif msg_type == 5:  # Static and Voyage Related Data
            vessel_doc = {
                'mmsi': mmsi,
                'name': decoded.get('shipname', '').strip(),
                'callsign': decoded.get('callsign', '').strip(),
                'imo': decoded.get('imo'),
                'ship_type': decoded.get('shiptype'),
                'ship_type_text': get_ship_type_text(decoded.get('shiptype', 0)),
                'dimension_a': decoded.get('to_bow'),
                'dimension_b': decoded.get('to_stern'),
                'dimension_c': decoded.get('to_port'),
                'dimension_d': decoded.get('to_starboard'),
                'destination': decoded.get('destination', '').strip(),
                'eta': str(decoded.get('eta', '')),
                'last_seen': timestamp.isoformat(),
                'country': get_mmsi_country(mmsi)
            }
            
            await db.vessels.update_one(
                {'mmsi': mmsi},
                {
                    '$set': vessel_doc,
                    '$addToSet': {'source_ids': source_id}
                },
                upsert=True
            )
            
            # Broadcast to WebSocket clients
            await manager.broadcast({
                'type': 'vessel_info',
                'data': vessel_doc
            })
        
        elif msg_type == 18:  # Standard Class B Position Report
            original_lat = decoded.get('lat')
            original_lon = decoded.get('lon')
            
            # Validate position
            position_is_valid = is_valid_position(original_lat, original_lon)
            
            # Determine display coordinates
            if position_is_valid:
                display_lat = original_lat
                display_lon = original_lon
            else:
                # Try to get last valid position
                last_valid = await get_last_valid_position(mmsi)
                if last_valid:
                    display_lat = last_valid['lat']
                    display_lon = last_valid['lon']
                else:
                    # No previous valid position, leave as None for now
                    display_lat = None
                    display_lon = None
            
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': original_lat,
                'lon': original_lon,
                'display_lat': display_lat,
                'display_lon': display_lon,
                'position_valid': position_is_valid,
                'speed': decoded.get('speed'),
                'course': decoded.get('course'),
                'heading': decoded.get('heading'),
                'source_id': source_id,
                'is_vdo': is_vdo,
                'repeat_indicator': decoded.get('repeat', 0)
            }
            await db.positions.insert_one(position_doc)
            
            # If this is the first valid position, backfill any previous invalid positions
            if position_is_valid:
                await backfill_invalid_positions(mmsi, display_lat, display_lon)
            
            # Only update vessel if we have valid display coordinates
            if display_lat is not None and display_lon is not None:
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_position': position_doc, 
                            'last_seen': timestamp.isoformat(),
                            'country': get_mmsi_country(mmsi)
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
                
                await manager.broadcast({
                    'type': 'position',
                    'data': position_doc
                })
            else:
                # No valid position yet, just update metadata
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_seen': timestamp.isoformat(),
                            'country': get_mmsi_country(mmsi)
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
        
        elif msg_type == 24:  # Static Data Report
            vessel_doc = {
                'mmsi': mmsi,
                'last_seen': timestamp.isoformat(),
                'country': get_mmsi_country(mmsi)
            }
            
            if 'shipname' in decoded:
                vessel_doc['name'] = decoded['shipname'].strip()
            if 'callsign' in decoded:
                vessel_doc['callsign'] = decoded['callsign'].strip()
            if 'shiptype' in decoded:
                vessel_doc['ship_type'] = decoded['shiptype']
                vessel_doc['ship_type_text'] = get_ship_type_text(decoded['shiptype'])
            
            await db.vessels.update_one(
                {'mmsi': mmsi},
                {
                    '$set': vessel_doc,
                    '$addToSet': {'source_ids': source_id}
                },
                upsert=True
            )
            
            await manager.broadcast({
                'type': 'vessel_info',
                'data': vessel_doc
            })
        
        elif msg_type == 21:  # Aid to Navigation (AtoN) Report
            original_lat = decoded.get('lat')
            original_lon = decoded.get('lon')
            
            # Validate position
            position_is_valid = is_valid_position(original_lat, original_lon)
            
            # Determine display coordinates
            if position_is_valid:
                display_lat = original_lat
                display_lon = original_lon
            else:
                # Try to get last valid position
                last_valid = await get_last_valid_position(mmsi)
                if last_valid:
                    display_lat = last_valid['lat']
                    display_lon = last_valid['lon']
                else:
                    display_lat = None
                    display_lon = None
            
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': original_lat,
                'lon': original_lon,
                'display_lat': display_lat,
                'display_lon': display_lon,
                'position_valid': position_is_valid,
                'source_id': source_id,
                'is_vdo': is_vdo,
                'is_aton': True,
                'aton_type': decoded.get('aid_type'),
                'aton_name': decoded.get('name', '').strip() if decoded.get('name') else '',
                'off_position': decoded.get('off_position', False),
                'virtual_aid': decoded.get('virtual_aid', False),
                'repeat_indicator': decoded.get('repeat', 0)
            }
            
            # Always store the position
            await db.positions.insert_one(position_doc)
            
            # If this is the first valid position, backfill any previous invalid positions
            if position_is_valid:
                await backfill_invalid_positions(mmsi, display_lat, display_lon)
            
            # Get position count
            pos_count = await db.positions.count_documents({'mmsi': mmsi})
            
            # Only update vessel if we have valid display coordinates
            if display_lat is not None and display_lon is not None:
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_position': position_doc,
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi),
                            'is_aton': True,
                            'name': position_doc['aton_name']
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
                
                # Broadcast to WebSocket clients
                await manager.broadcast({
                    'type': 'position',
                    'data': position_doc
                })
            else:
                # No valid position yet, just update metadata
                await db.vessels.update_one(
                    {'mmsi': mmsi},
                    {
                        '$set': {
                            'last_seen': timestamp.isoformat(),
                            'position_count': pos_count,
                            'country': get_mmsi_country(mmsi),
                            'is_aton': True,
                            'name': position_doc['aton_name']
                        },
                        '$addToSet': {'source_ids': source_id}
                    },
                    upsert=True
                )
        
        # Update source statistics
        if source_id:
            target_count = await db.vessels.count_documents({'source_ids': source_id})
            await db.sources.update_one(
                {'source_id': source_id},
                {
                    '$inc': {'message_count': 1},
                    '$set': {
                        'target_count': target_count,
                        'last_message': timestamp.isoformat()
                    }
                }
            )
        
        logger.info(f"Processed AIS message type {msg_type} for MMSI {mmsi}")
        
    except Exception as e:
        logger.error(f"Error processing AIS message: {e}")

# Routes
@api_router.get("/")
async def root():
    return {"message": "AIS/NMEA Tracking System API"}

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), background_tasks: BackgroundTasks = None):
    """Upload and process AIS log file"""
    try:
        # Check for duplicate filename
        existing = await db.sources.find_one({
            'source_type': 'file',
            'config.filename': file.filename
        })
        
        if existing:
            logger.warning(f"Duplicate file upload rejected: {file.filename}")
            raise HTTPException(
                status_code=409,
                detail=f"File '{file.filename}' has already been uploaded. Please delete the existing source first if you want to re-upload."
            )
        
        # Create source record
        source_id = str(uuid.uuid4())
        source_doc = {
            'source_id': source_id,
            'source_type': 'file',
            'name': file.filename,
            'config': {'filename': file.filename},
            'status': 'active',
            'created_at': datetime.now(timezone.utc).isoformat(),
            'message_count': 0
        }
        await db.sources.insert_one(source_doc)
        
        content = await file.read()
        lines = content.decode('utf-8', errors='ignore').split('\n')
        
        processed = 0
        errors = 0
        
        for line in lines:
            line = line.strip()
            if line and (line.startswith('!') or line.startswith('$')):
                try:
                    await process_ais_message(line, source=f'file:{file.filename}', source_id=source_id)
                    processed += 1
                except Exception as e:
                    errors += 1
                    logger.error(f"Error processing line: {e}")
        
        # Update source with message count
        await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {
                'message_count': processed,
                'last_message': datetime.now(timezone.utc).isoformat()
            }}
        )
        
        return {
            'status': 'success',
            'source_id': source_id,
            'filename': file.filename,
            'processed': processed,
            'errors': errors
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/stream/start")
async def start_stream(config: StreamConfig, background_tasks: BackgroundTasks):
    """Start TCP/UDP/Serial stream"""
    
    # Check for duplicates
    duplicate_query = {'source_type': config.stream_type}
    
    if config.stream_type in ['tcp', 'udp']:
        duplicate_query['config.host'] = config.host
        duplicate_query['config.port'] = config.port
        duplicate_name = f"{config.stream_type.upper()}: {config.host}:{config.port}"
    elif config.stream_type == 'serial':
        duplicate_query['config.serial_port'] = config.serial_port
        duplicate_name = f"SERIAL: {config.serial_port}"
    
    existing = await db.sources.find_one(duplicate_query)
    if existing:
        logger.warning(f"Duplicate source rejected: {duplicate_name}")
        raise HTTPException(
            status_code=409, 
            detail=f"Source already exists: {duplicate_name}. Please enable the existing source instead of adding a duplicate."
        )
    
    source_id = str(uuid.uuid4())
    
    # Create source record
    source_name = f"{config.stream_type.upper()}: "
    if config.stream_type in ['tcp', 'udp']:
        source_name += f"{config.host}:{config.port}"
    else:
        source_name += config.serial_port
    
    source_doc = {
        'source_id': source_id,
        'source_type': config.stream_type,
        'name': source_name,
        'config': config.dict(),
        'status': 'active',
        'created_at': datetime.now(timezone.utc).isoformat(),
        'message_count': 0
    }
    await db.sources.insert_one(source_doc)
    
    stream_id = source_id
    
    # Async helper to process message and update stats
    async def process_stream_message(raw_msg: str, source_str: str, sid: str):
        await process_ais_message(raw_msg, source=source_str, source_id=sid)
        await db.sources.update_one(
            {'source_id': sid},
            {
                '$inc': {'message_count': 1},
                '$set': {'last_message': datetime.now(timezone.utc).isoformat()}
            }
        )
    
    def tcp_stream_handler():
        try:
            logger.info(f"Connecting to TCP stream: {config.host}:{config.port}")
            conn = TCPConnection(config.host, port=config.port)
            
            logger.info(f"TCP stream connected: {config.host}:{config.port}")
            
            for msg in conn:
                if source_id not in active_streams:
                    logger.info(f"TCP stream {source_id} stopped")
                    break
                try:
                    # msg is an AISSentence object with .raw attribute (bytes)
                    raw_msg = msg.raw.decode('utf-8', errors='ignore')
                    
                    # Schedule the async function on the main event loop
                    if main_event_loop:
                        future = asyncio.run_coroutine_threadsafe(
                            process_stream_message(raw_msg, f'tcp:{config.host}:{config.port}', source_id),
                            main_event_loop
                        )
                        # Wait for completion (with timeout to prevent blocking)
                        try:
                            future.result(timeout=2)
                        except Exception as e:
                            logger.error(f"Error processing TCP message: {e}")
                    
                except Exception as e:
                    logger.error(f"Error in TCP handler: {e}")
                    
        except Exception as e:
            logger.error(f"TCP stream error for {config.host}:{config.port} - {e}")
    
    def udp_stream_handler():
        try:
            logger.info(f"Starting UDP receiver: {config.host}:{config.port}")
            receiver = UDPReceiver(config.host, port=config.port)
            
            logger.info(f"UDP receiver started: {config.host}:{config.port}")
            
            for msg in receiver:
                if source_id not in active_streams:
                    logger.info(f"UDP stream {source_id} stopped")
                    break
                try:
                    raw_msg = msg.raw.decode('utf-8', errors='ignore')
                    
                    # Schedule the async function on the main event loop
                    if main_event_loop:
                        future = asyncio.run_coroutine_threadsafe(
                            process_stream_message(raw_msg, f'udp:{config.host}:{config.port}', source_id),
                            main_event_loop
                        )
                        # Wait for completion (with timeout to prevent blocking)
                        try:
                            future.result(timeout=2)
                        except Exception as e:
                            logger.error(f"Error processing UDP message: {e}")
                    
                except Exception as e:
                    logger.error(f"Error in UDP handler: {e}")
                    
        except Exception as e:
            logger.error(f"UDP stream error: {e}")
    
    def serial_stream_handler():
        try:
            logger.info(f"Opening serial port: {config.serial_port} at {config.baudrate} baud")
            ser = serial.Serial(config.serial_port, config.baudrate, timeout=1)
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            logger.info(f"Serial port opened: {config.serial_port}")
            
            while source_id in active_streams:
                try:
                    line = ser.readline().decode('utf-8', errors='ignore').strip()
                    if line:
                        loop.run_until_complete(process_ais_message(line, source=f'serial:{config.serial_port}', source_id=source_id))
                        loop.run_until_complete(db.sources.update_one(
                            {'source_id': source_id},
                            {
                                '$inc': {'message_count': 1},
                                '$set': {'last_message': datetime.now(timezone.utc).isoformat()}
                            }
                        ))
                except Exception as e:
                    logger.error(f"Serial read error: {e}")
            ser.close()
        except Exception as e:
            logger.error(f"Serial stream error: {e}")
    
    # Start appropriate stream
    if config.stream_type == 'tcp':
        active_streams[stream_id] = 'tcp'
        thread = threading.Thread(target=tcp_stream_handler, daemon=True)
        thread.start()
    elif config.stream_type == 'udp':
        active_streams[stream_id] = 'udp'
        thread = threading.Thread(target=udp_stream_handler, daemon=True)
        thread.start()
    elif config.stream_type == 'serial':
        active_streams[stream_id] = 'serial'
        thread = threading.Thread(target=serial_stream_handler, daemon=True)
        thread.start()
    
    return {'status': 'started', 'stream_id': stream_id}

@api_router.post("/stream/stop/{stream_id}")
async def stop_stream(stream_id: str):
    """Stop active stream"""
    if stream_id in active_streams:
        del active_streams[stream_id]
        return {'status': 'stopped'}
    return {'status': 'not_found'}

@api_router.get("/stream/active")
async def get_active_streams():
    """Get list of active streams"""
    return {'streams': active_streams}

@api_router.get("/sources")
async def get_sources():
    """Get all data sources"""
    try:
        sources = await db.sources.find().sort('created_at', -1).to_list(100)
        serialized_sources = [serialize_doc(s) for s in sources]
        return {'sources': serialized_sources}
    except Exception as e:
        logger.error(f"Error loading sources: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/sources/{source_id}/toggle")
async def toggle_source(source_id: str):
    """Enable/disable a data source"""
    try:
        source = await db.sources.find_one({'source_id': source_id})
        if not source:
            raise HTTPException(status_code=404, detail="Source not found")
        
        new_status = 'inactive' if source['status'] == 'active' else 'active'
        
        # If disabling a stream, stop it
        if new_status == 'inactive' and source_id in active_streams:
            logger.info(f"Stopping stream {source_id}")
            del active_streams[source_id]
        
        # If enabling a stream, restart it
        if new_status == 'active' and source['source_type'] in ['tcp', 'udp', 'serial']:
            logger.info(f"Restarting stream {source_id}")
            config = source.get('config', {})
            
            # Mark as active in streams dict
            active_streams[source_id] = True
            
            # Start stream handler in background thread
            if source['source_type'] == 'tcp':
                def restart_tcp():
                    try:
                        from pyais.stream import TCPConnection
                        conn = TCPConnection(config['host'], port=config['port'])
                        logger.info(f"TCP stream {source_id} reconnected")
                        
                        for msg in conn:
                            if source_id not in active_streams:
                                break
                            try:
                                raw_msg = msg.raw.decode('utf-8', errors='ignore')
                                if main_event_loop:
                                    async def process():
                                        await process_ais_message(raw_msg, source=f"tcp:{config['host']}:{config['port']}", source_id=source_id)
                                    future = asyncio.run_coroutine_threadsafe(process(), main_event_loop)
                                    future.result(timeout=2)
                            except Exception as e:
                                logger.error(f"Error processing TCP message: {e}")
                    except Exception as e:
                        logger.error(f"TCP stream error: {e}")
                
                thread = threading.Thread(target=restart_tcp, daemon=True)
                thread.start()
                
            elif source['source_type'] == 'udp':
                def restart_udp():
                    try:
                        from pyais.stream import UDPReceiver
                        receiver = UDPReceiver(config['host'], port=config['port'])
                        logger.info(f"UDP stream {source_id} reconnected")
                        
                        for msg in receiver:
                            if source_id not in active_streams:
                                break
                            try:
                                raw_msg = msg.raw.decode('utf-8', errors='ignore')
                                if main_event_loop:
                                    async def process():
                                        await process_ais_message(raw_msg, source=f"udp:{config['host']}:{config['port']}", source_id=source_id)
                                    future = asyncio.run_coroutine_threadsafe(process(), main_event_loop)
                                    future.result(timeout=2)
                            except Exception as e:
                                logger.error(f"Error processing UDP message: {e}")
                    except Exception as e:
                        logger.error(f"UDP stream error: {e}")
                
                thread = threading.Thread(target=restart_udp, daemon=True)
                thread.start()
                
            elif source['source_type'] == 'serial':
                def restart_serial():
                    try:
                        from pyais.stream import SerialStream
                        stream = SerialStream(config['serial_port'], baudrate=config.get('baudrate', 9600))
                        logger.info(f"Serial stream {source_id} reconnected")
                        
                        for msg in stream:
                            if source_id not in active_streams:
                                break
                            try:
                                raw_msg = msg.raw.decode('utf-8', errors='ignore')
                                if main_event_loop:
                                    async def process():
                                        await process_ais_message(raw_msg, source=f"serial:{config['serial_port']}", source_id=source_id)
                                    future = asyncio.run_coroutine_threadsafe(process(), main_event_loop)
                                    future.result(timeout=2)
                            except Exception as e:
                                logger.error(f"Error processing serial message: {e}")
                    except Exception as e:
                        logger.error(f"Serial stream error: {e}")
                
                thread = threading.Thread(target=restart_serial, daemon=True)
                thread.start()
        
        await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'status': new_status}}
        )
        
        return {'status': new_status}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error toggling source: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/sources/{source_id}/spoof-limit")
async def update_spoof_limit(source_id: str, spoof_limit_km: float):
    """Update spoof limit for a data source"""
    try:
        result = await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'spoof_limit_km': spoof_limit_km}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Source not found")
        
        return {'status': 'updated', 'spoof_limit_km': spoof_limit_km}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating spoof limit: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sources/{source_id}/pause")
async def pause_source(source_id: str):
    """Pause a streaming source (stops processing but keeps connection)"""
    try:
        source = await db.sources.find_one({'source_id': source_id})
        if not source:
            raise HTTPException(status_code=404, detail="Source not found")
        
        await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'is_paused': True}}
        )
        
        logger.info(f"Source {source_id} paused")
        return {'status': 'paused'}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error pausing source: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sources/{source_id}/resume")
async def resume_source(source_id: str):
    """Resume a paused streaming source"""
    try:
        source = await db.sources.find_one({'source_id': source_id})
        if not source:
            raise HTTPException(status_code=404, detail="Source not found")
        
        await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'is_paused': False}}
        )
        
        logger.info(f"Source {source_id} resumed")
        return {'status': 'resumed'}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error resuming source: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/sources/{source_id}/message-limit")
async def update_message_limit(source_id: str, message_limit: int):
    """Update message limit for a data source"""
    try:
        if message_limit < 10:
            raise HTTPException(status_code=400, detail="Message limit must be at least 10")
        
        result = await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'message_limit': message_limit}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Source not found")
        
        logger.info(f"Source {source_id} message limit updated to {message_limit}")
        return {'status': 'updated', 'message_limit': message_limit}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating message limit: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.patch("/sources/{source_id}/target-limit")
async def update_target_limit(source_id: str, target_limit: int):
    """Update target display limit for a data source (0 = unlimited)"""
    try:
        if target_limit < 0:
            raise HTTPException(status_code=400, detail="Target limit must be 0 (unlimited) or positive")
        
        result = await db.sources.update_one(
            {'source_id': source_id},
            {'$set': {'target_limit': target_limit}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Source not found")
        
        logger.info(f"Source {source_id} target limit updated to {target_limit}")
        return {'status': 'updated', 'target_limit': target_limit}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating target limit: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/sources/{source_id}")
async def delete_source(source_id: str, delete_data: bool = False):
    """Remove a data source and optionally its associated data"""
    try:
        # Stop stream if active
        if source_id in active_streams:
            del active_streams[source_id]
        
        # If requested, delete all associated data
        if delete_data:
            # Delete messages
            messages_result = await db.messages.delete_many({'source_id': source_id})
            
            # Delete positions
            positions_result = await db.positions.delete_many({'source_id': source_id})
            
            # Remove source_id from vessels' source_ids array
            await db.vessels.update_many(
                {'source_ids': source_id},
                {'$pull': {'source_ids': source_id}}
            )
            
            # Delete vessels that have no more sources
            vessels_result = await db.vessels.delete_many({'source_ids': {'$size': 0}})
            
            logger.info(f"Deleted data for source {source_id}: {messages_result.deleted_count} messages, {positions_result.deleted_count} positions, {vessels_result.deleted_count} vessels")
        
        # Delete the source itself
        result = await db.sources.delete_one({'source_id': source_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Source not found")
        
        return {
            'status': 'deleted',
            'data_deleted': delete_data,
            'messages_deleted': messages_result.deleted_count if delete_data else 0,
            'positions_deleted': positions_result.deleted_count if delete_data else 0,
            'vessels_deleted': vessels_result.deleted_count if delete_data else 0
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting source: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/sources/disable-all")
async def disable_all_sources():
    """Disable all data sources"""
    try:
        # Stop all active streams
        for stream_id in list(active_streams.keys()):
            del active_streams[stream_id]
        
        # Disable all sources
        await db.sources.update_many({}, {'$set': {'status': 'inactive'}})
        
        return {'status': 'all_disabled', 'count': await db.sources.count_documents({})}
    except Exception as e:
        logger.error(f"Error disabling all sources: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/vessels/active")
async def get_active_vessels(limit: int = 5000, skip: int = 0):
    """Get vessels from active sources with per-source VDO spoof detection"""
    try:
        import math
        
        # Get active sources with their spoof limits
        active_sources = await db.sources.find({'status': 'active'}).to_list(100)
        active_source_ids = [s['source_id'] for s in active_sources]
        
        if not active_source_ids:
            return {'vessels': [], 'total': 0, 'vdo_data': []}
        
        # Collect vessels per source with target limiting
        all_vessels_dict = {}  # Use dict to dedupe by MMSI
        
        for source in active_sources:
            source_id = source['source_id']
            target_limit = source.get('target_limit', 0)  # 0 = unlimited
            
            # Get vessels from this source, sorted by most recent
            query = {'source_ids': source_id}
            
            if target_limit > 0:
                # Limited: get only N most recent targets
                source_vessels = await db.vessels.find(query).sort('last_seen', -1).limit(target_limit).to_list(target_limit)
            else:
                # Unlimited: get all targets from this source
                source_vessels = await db.vessels.find(query).sort('last_seen', -1).to_list(10000)
            
            # Add to combined dict (dedupe by MMSI, keeping most recent last_seen)
            for vessel in source_vessels:
                mmsi = vessel['mmsi']
                if mmsi not in all_vessels_dict or vessel['last_seen'] > all_vessels_dict[mmsi]['last_seen']:
                    all_vessels_dict[mmsi] = vessel
        
        # Convert dict back to list and sort by last_seen
        vessels = sorted(all_vessels_dict.values(), key=lambda v: v.get('last_seen', ''), reverse=True)
        
        # Apply pagination
        total = len(vessels)
        vessels = vessels[skip:skip + limit]
        
        # Process VDO positions per source
        vdo_data_list = []
        
        for source in active_sources:
            source_id = source['source_id']
            spoof_limit_km = source.get('spoof_limit_km', 500.0)
            
            # Get ALL base station positions for this source (both own VDO and received VDM Type 4)
            # Query for base stations: is_base_station flag OR is_vdo flag
            vdo_positions = await db.positions.find({
                '$or': [
                    {'is_base_station': True},
                    {'is_vdo': True}
                ],
                'source_id': source_id,
                'display_lat': {'$exists': True, '$ne': None},
                'display_lon': {'$exists': True, '$ne': None}
            }).to_list(100)
            
            for vdo_pos in vdo_positions:
                # Use display coordinates for plotting
                vdo_lat = vdo_pos.get('display_lat')
                vdo_lon = vdo_pos.get('display_lon')
                vdo_mmsi = vdo_pos.get('mmsi')
                
                if not vdo_lat or not vdo_lon:
                    continue
                
                # Get VDM positions from SAME source only
                # Exclude VDO messages and only get VDMs with repeat_indicator <= 0
                # Only use positions with valid display coordinates
                vdm_positions = await db.positions.find({
                    'is_vdo': {'$ne': True},
                    'source_id': source_id,
                    'repeat_indicator': {'$lte': 0},  # Only direct messages, not repeated
                    'display_lat': {'$exists': True, '$ne': None},
                    'display_lon': {'$exists': True, '$ne': None}
                }).to_list(10000)
                
                # Find furthest VDM within spoof limit (with repeat_indicator <= 0)
                max_distance_within_limit = 0
                
                for vdm_pos in vdm_positions:
                    vdm_lat = vdm_pos.get('display_lat')
                    vdm_lon = vdm_pos.get('display_lon')
                    repeat_ind = vdm_pos.get('repeat_indicator', 0)
                    
                    # Double-check: only valid VDMs (repeat_indicator <= 0)
                    if vdm_lat and vdm_lon and repeat_ind <= 0:
                        # Calculate distance in km
                        lat1, lon1 = math.radians(vdo_lat), math.radians(vdo_lon)
                        lat2, lon2 = math.radians(vdm_lat), math.radians(vdm_lon)
                        
                        dlat = lat2 - lat1
                        dlon = lon2 - lon1
                        
                        a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
                        c = 2 * math.asin(math.sqrt(a))
                        distance_km = 6371 * c
                        
                        # Only consider VDMs within spoof limit AND with repeat_indicator <= 0
                        if distance_km <= spoof_limit_km:
                            max_distance_within_limit = max(max_distance_within_limit, distance_km)
                
                # Determine if this is own base station (VDO) or received (VDM Type 4)
                is_own_base_station = vdo_pos.get('is_vdo', False)
                
                vdo_data_list.append({
                    'mmsi': vdo_mmsi,
                    'lat': vdo_lat,
                    'lon': vdo_lon,
                    'radius_km': max_distance_within_limit,
                    'spoof_limit_km': spoof_limit_km,
                    'source_id': source_id,
                    'source_name': source['name'],
                    'timestamp': vdo_pos.get('timestamp'),
                    'is_own': is_own_base_station  # True = own/VDO, False = received/VDM
                })
        
        # Count how many sources each base station MMSI appears in
        base_station_source_counts = {}
        for vdo in vdo_data_list:
            mmsi = vdo['mmsi']
            if mmsi not in base_station_source_counts:
                base_station_source_counts[mmsi] = set()
            base_station_source_counts[mmsi].add(vdo['source_id'])
        
        # Add multi_source flag to VDO data
        for vdo in vdo_data_list:
            mmsi = vdo['mmsi']
            vdo['source_count'] = len(base_station_source_counts.get(mmsi, set()))
            vdo['multi_source'] = vdo['source_count'] > 1
        
        logger.info(f"Found {len(vessels)}/{total} vessels, {len(vdo_data_list)} VDO positions")
        
        serialized_vessels = [serialize_doc(v) for v in vessels]
        return {
            'vessels': serialized_vessels,
            'total': total,
            'vdo_data': vdo_data_list
        }
    except Exception as e:
        logger.error(f"Error loading active vessels: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/database/clear")
async def clear_database():
    """Clear all vessel, position, and message data (keep sources)"""
    try:
        vessels_deleted = await db.vessels.delete_many({})
        positions_deleted = await db.positions.delete_many({})
        messages_deleted = await db.messages.delete_many({})
        
        logger.info(f"Database cleared: {vessels_deleted.deleted_count} vessels, {positions_deleted.deleted_count} positions, {messages_deleted.deleted_count} messages")
        
        return {
            'status': 'cleared',
            'vessels_deleted': vessels_deleted.deleted_count,
            'positions_deleted': positions_deleted.deleted_count,
            'messages_deleted': messages_deleted.deleted_count
        }
    except Exception as e:
        logger.error(f"Error clearing database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/database/update-spoof-limits")
async def update_spoof_limits():
    """Update all sources with old 50km spoof limit to new 500km default"""
    try:
        result = await db.sources.update_many(
            {'spoof_limit_km': 50.0},
            {'$set': {'spoof_limit_km': 500.0}}
        )
        
        logger.info(f"Updated {result.modified_count} sources from 50km to 500km spoof limit")
        
        return {
            'status': 'updated',
            'sources_updated': result.modified_count
        }
    except Exception as e:
        logger.error(f"Error updating spoof limits: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/status")
async def get_status():
    """Get system status and statistics"""
    try:
        # Count vessels, messages, positions
        vessel_count = await db.vessels.count_documents({})
        message_count = await db.messages.count_documents({})
        position_count = await db.positions.count_documents({})
        source_count = await db.sources.count_documents({})
        
        # Get sources with their message rates
        sources = await db.sources.find().to_list(1000)
        source_stats = []
        
        for source in sources:
            stats = {
                'name': source.get('name'),
                'type': source.get('source_type'),
                'status': source.get('status'),
                'message_count': source.get('message_count', 0),
                'target_count': source.get('target_count', 0),
                'fragment_count': source.get('fragment_count', 0),
                'created_at': source.get('created_at'),
                'last_message': source.get('last_message')
            }
            
            # Calculate messages per second for active streaming sources
            if source.get('status') == 'active' and source.get('source_type') in ['tcp', 'udp', 'serial']:
                created = datetime.fromisoformat(source.get('created_at'))
                now = datetime.now(timezone.utc)
                elapsed_seconds = (now - created).total_seconds()
                if elapsed_seconds > 0:
                    stats['messages_per_second'] = round(source.get('message_count', 0) / elapsed_seconds, 2)
                else:
                    stats['messages_per_second'] = 0
            
            source_stats.append(stats)
        
        return {
            'vessels': vessel_count,
            'messages': message_count,
            'positions': position_count,
            'sources': source_count,
            'source_stats': source_stats
        }
    except Exception as e:
        logger.error(f"Error getting status: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/export/xlsx")
async def export_xlsx():
    """Export ALL AIS data to Excel file with comprehensive details"""
    try:
        logger.info("Starting comprehensive XLSX export...")
        
        # Create workbook
        wb = Workbook()
        
        # Create source lookup first
        sources = await db.sources.find().to_list(1000)
        source_lookup = {s['source_id']: s.get('name', 'Unknown') for s in sources}
        logger.info(f"Loaded {len(sources)} sources")
        
        # Sheet 1: All Positions (Historical)
        ws_positions = wb.active
        ws_positions.title = "All Positions"
        ws_positions.append([
            'MMSI', 'Timestamp', 'Message Type', 'Original Lat', 'Original Lon', 
            'Display Lat', 'Display Lon', 'Position Valid', 'Backfilled',
            'Speed (knots)', 'Course', 'Heading', 'Navigation Status', 
            'Accuracy', 'ROT', 'RAIM', 'EPFD',
            'Is VDO', 'Is Base Station', 'Repeat Indicator', 
            'Source ID', 'Source Name'
        ])
        
        # Get positions count first
        position_count = await db.positions.count_documents({})
        logger.info(f"Total positions in database: {position_count}")
        
        # Limit to 500k positions to prevent memory issues
        max_positions = min(position_count, 500000)
        positions = await db.positions.find().sort('timestamp', -1).limit(max_positions).to_list(max_positions)
        logger.info(f"Exporting {len(positions)} positions...")
        
        for idx, pos in enumerate(positions):
            try:
                ws_positions.append([
                    pos.get('mmsi'),
                    pos.get('timestamp'),
                    pos.get('message_type'),
                    pos.get('lat'),  # Original latitude
                    pos.get('lon'),  # Original longitude
                    pos.get('display_lat'),  # Display latitude
                    pos.get('display_lon'),  # Display longitude
                    pos.get('position_valid'),
                    pos.get('backfilled', False),
                    pos.get('speed'),
                    pos.get('course'),
                    pos.get('heading'),
                    pos.get('nav_status'),
                    pos.get('accuracy'),
                    pos.get('rot'),
                    pos.get('raim'),
                    pos.get('epfd'),
                    pos.get('is_vdo', False),
                    pos.get('is_base_station', False),
                    pos.get('repeat_indicator'),
                    pos.get('source_id', '')[:8] + '...' if pos.get('source_id') else '',
                    source_lookup.get(pos.get('source_id'), 'Unknown')
                ])
                
                # Log progress every 10000 rows
                if (idx + 1) % 10000 == 0:
                    logger.info(f"Exported {idx + 1}/{len(positions)} positions...")
            except Exception as e:
                logger.error(f"Error exporting position {idx}: {e}")
                continue
        
        # Sheet 2: All Messages (Raw + Decoded)
        ws_messages = wb.create_sheet("All Messages")
        ws_messages.append([
            'MMSI', 'Timestamp', 'Message Type', 'Raw NMEA', 
            'Is VDO', 'Repeat Indicator', 'Source ID', 'Source Name',
            'Decoded - Ship Name', 'Decoded - Call Sign', 'Decoded - IMO',
            'Decoded - Ship Type', 'Decoded - Destination', 'Decoded - ETA',
            'Decoded - Dimensions (A/B/C/D)', 'Decoded - Draught'
        ])
        
        # Get messages count first
        message_count = await db.messages.count_documents({})
        logger.info(f"Total messages in database: {message_count}")
        
        # Limit to 500k messages to prevent memory issues
        max_messages = min(message_count, 500000)
        messages = await db.messages.find().sort('timestamp', -1).limit(max_messages).to_list(max_messages)
        logger.info(f"Exporting {len(messages)} messages...")
        
        for idx, msg in enumerate(messages):
            try:
                decoded = msg.get('decoded', {})
                dimensions = f"{decoded.get('to_bow', '')}/{decoded.get('to_stern', '')}/{decoded.get('to_port', '')}/{decoded.get('to_starboard', '')}"
                
                # Safe string handling for decoded fields
                shipname = decoded.get('shipname')
                callsign = decoded.get('callsign')
                destination = decoded.get('destination')
                
                ws_messages.append([
                    msg.get('mmsi'),
                    msg.get('timestamp'),
                    msg.get('message_type'),
                    msg.get('raw'),  # Original NMEA sentence
                    msg.get('is_vdo', False),
                    msg.get('repeat_indicator'),
                    msg.get('source_id', '')[:8] + '...' if msg.get('source_id') else '',
                    source_lookup.get(msg.get('source_id'), 'Unknown'),
                    shipname.strip() if shipname else '',
                    callsign.strip() if callsign else '',
                    decoded.get('imo'),
                    decoded.get('shiptype'),
                    destination.strip() if destination else '',
                    str(decoded.get('eta', '')),
                    dimensions,
                    decoded.get('draught')
                ])
                
                # Log progress every 10000 rows
                if (idx + 1) % 10000 == 0:
                    logger.info(f"Exported {idx + 1}/{len(messages)} messages...")
            except Exception as e:
                logger.error(f"Error exporting message {idx}: {e}")
                continue
        
        # Sheet 3: Vessels Summary
        ws_vessels = wb.create_sheet("Vessels Summary")
        ws_vessels.append([
            'MMSI', 'Ship Name', 'Call Sign', 'IMO', 'Ship Type', 'Ship Type Text',
            'Country', 'Is Base Station', 'Destination', 'ETA', 
            'Dimensions (A/B/C/D)', 'Position Count', 'Last Seen', 
            'Current Lat', 'Current Lon', 'Current Speed', 'Current Course', 'Current Heading',
            'Sources'
        ])
        
        vessels = await db.vessels.find().to_list(10000)
        logger.info(f"Exporting {len(vessels)} vessels...")
        
        for vessel in vessels:
            last_pos = vessel.get('last_position', {})
            dimensions = f"{vessel.get('dimension_a', '')}/{vessel.get('dimension_b', '')}/{vessel.get('dimension_c', '')}/{vessel.get('dimension_d', '')}"
            
            ws_vessels.append([
                vessel.get('mmsi'),
                vessel.get('name', ''),
                vessel.get('callsign', ''),
                vessel.get('imo'),
                vessel.get('ship_type'),
                vessel.get('ship_type_text', ''),
                vessel.get('country', ''),
                vessel.get('is_base_station', False),
                vessel.get('destination', ''),
                vessel.get('eta', ''),
                dimensions,
                vessel.get('position_count', 0),
                vessel.get('last_seen'),
                last_pos.get('display_lat') or last_pos.get('lat'),
                last_pos.get('display_lon') or last_pos.get('lon'),
                last_pos.get('speed'),
                last_pos.get('course'),
                last_pos.get('heading'),
                ', '.join(vessel.get('source_ids', [])[:3])  # First 3 sources
            ])
        
        # Sheet 4: Sources
        ws_sources = wb.create_sheet("Sources")
        ws_sources.append([
            'Source ID', 'Name', 'Type', 'Status', 'Host', 'Port',
            'Message Count', 'Target Count', 'Fragment Count', 
            'Spoof Limit (km)', 'Message Limit', 'Is Paused',
            'Created', 'Last Message'
        ])
        
        for source in sources:
            ws_sources.append([
                source.get('source_id', '')[:12] + '...',
                source.get('name'),
                source.get('source_type'),
                source.get('status'),
                source.get('host', ''),
                source.get('port', ''),
                source.get('message_count', 0),
                source.get('target_count', 0),
                source.get('fragment_count', 0),
                source.get('spoof_limit_km', 500),
                source.get('message_limit', 500),
                source.get('is_paused', False),
                source.get('created_at'),
                source.get('last_message')
            ])
        
        logger.info("Saving workbook...")
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info("Export complete!")
        
        # Return as downloadable file
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=ais_complete_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'}
        )
    except Exception as e:
        logger.error(f"Error exporting xlsx: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/serial/ports")
async def list_serial_ports():
    """List available serial ports"""
    ports = serial.tools.list_ports.comports()
    return {'ports': [{'device': p.device, 'description': p.description} for p in ports]}

@api_router.get("/vdo/positions")
async def get_vdo_positions():
    """Get all VDO (own ship) positions with spoof detection radius"""
    try:
        import math
        
        # Get all VDO positions
        vdo_positions = await db.positions.find({'is_vdo': True}).to_list(1000)
        
        result = []
        
        for vdo_pos in vdo_positions:
            if not vdo_pos.get('lat') or not vdo_pos.get('lon'):
                continue
                
            vdo_lat = vdo_pos['lat']
            vdo_lon = vdo_pos['lon']
            
            # Get all VDM positions with repeat_indicator <= 0
            vdm_positions = await db.positions.find({
                'is_vdo': {'$ne': True},
                'repeat_indicator': {'$lte': 0},
                'lat': {'$exists': True, '$ne': None},
                'lon': {'$exists': True, '$ne': None}
            }).to_list(10000)
            
            max_distance = 0
            
            # Calculate distance to furthest VDM target
            for vdm_pos in vdm_positions:
                vdm_lat = vdm_pos.get('lat')
                vdm_lon = vdm_pos.get('lon')
                
                if vdm_lat and vdm_lon:
                    # Haversine formula for distance
                    lat1, lon1 = math.radians(vdo_lat), math.radians(vdo_lon)
                    lat2, lon2 = math.radians(vdm_lat), math.radians(vdm_lon)
                    
                    dlat = lat2 - lat1
                    dlon = lon2 - lon1
                    
                    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
                    c = 2 * math.asin(math.sqrt(a))
                    distance_km = 6371 * c  # Earth radius in km
                    distance_nm = distance_km * 0.539957  # Convert to nautical miles
                    
                    max_distance = max(max_distance, distance_nm)
            
            result.append({
                'mmsi': vdo_pos.get('mmsi'),
                'lat': vdo_lat,
                'lon': vdo_lon,
                'radius_nm': max_distance,
                'radius_km': max_distance * 1.852,
                'timestamp': vdo_pos.get('timestamp')
            })
        
        return {'vdo_positions': result}
    except Exception as e:
        logger.error(f"Error getting VDO positions: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/vessels")
async def get_vessels(limit: int = 100):
    """Get all vessels"""
    try:
        vessels = await db.vessels.find().sort('last_seen', -1).limit(limit).to_list(limit)
        serialized_vessels = [serialize_doc(v) for v in vessels]
        return {'vessels': serialized_vessels}
    except Exception as e:
        logger.error(f"Error loading vessels: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load vessels: {str(e)}")

@api_router.get("/vessel/{mmsi}")
async def get_vessel(mmsi: str):
    """Get vessel details by MMSI - combines data from all sources"""
    try:
        vessel = await db.vessels.find_one({'mmsi': mmsi})
        if vessel:
            vessel = serialize_doc(vessel)
            
            # Get the absolute latest position across all sources
            latest_position = await db.positions.find_one(
                {'mmsi': mmsi},
                sort=[('timestamp', -1)]
            )
            if latest_position:
                vessel['last_position'] = serialize_doc(latest_position)
            
            # Get recent positions sorted by timestamp (newest first)
            positions = await db.positions.find({'mmsi': mmsi}).sort('timestamp', -1).limit(100).to_list(100)
            positions = [serialize_doc(p) for p in positions]
            
            vessel['track'] = positions
            return vessel
        raise HTTPException(status_code=404, detail="Vessel not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading vessel {mmsi}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load vessel: {str(e)}")

@api_router.post("/search")
async def search_vessels(query: SearchQuery):
    """Search vessels by MMSI, name, or callsign"""
    try:
        # Get active source IDs
        active_sources = await db.sources.find({'status': 'active'}).to_list(100)
        active_source_ids = [s['source_id'] for s in active_sources]
        
        if not active_source_ids:
            return {'vessels': []}
        
        filter_query = {'source_ids': {'$in': active_source_ids}}
        
        # Build OR query for MMSI, name, and callsign
        or_conditions = []
        if query.mmsi:
            or_conditions.append({'mmsi': {'$regex': query.mmsi, '$options': 'i'}})
        if query.vessel_name:
            or_conditions.append({'name': {'$regex': query.vessel_name, '$options': 'i'}})
            or_conditions.append({'callsign': {'$regex': query.vessel_name, '$options': 'i'}})
        if query.ship_type is not None:
            filter_query['ship_type'] = query.ship_type
        
        if or_conditions:
            filter_query['$or'] = or_conditions
        
        vessels = await db.vessels.find(filter_query).limit(500).to_list(500)
        serialized_vessels = [serialize_doc(v) for v in vessels]
        
        return {'vessels': serialized_vessels}
    except Exception as e:
        logger.error(f"Error searching vessels: {e}")
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")

@api_router.get("/positions/recent")
async def get_recent_positions(limit: int = 100):
    """Get recent positions for all vessels"""
    try:
        # Get unique MMSIs with their latest position
        pipeline = [
            {'$sort': {'timestamp': -1}},
            {'$group': {
                '_id': '$mmsi',
                'latest': {'$first': '$$ROOT'}
            }},
            {'$limit': limit}
        ]
        
        results = await db.positions.aggregate(pipeline).to_list(limit)
        positions = [serialize_doc(r['latest']) for r in results]
        
        return {'positions': positions}
    except Exception as e:
        logger.error(f"Error loading recent positions: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load positions: {str(e)}")

@api_router.get("/track/{mmsi}")
async def get_vessel_track(mmsi: str, limit: int = 10000):
    """Get vessel track history - properly sorted by timestamp across all sources"""
    try:
        # Get positions from ALL sources for this MMSI
        # Sort by timestamp descending (newest first), then reverse for trail drawing
        positions = await db.positions.find(
            {'mmsi': mmsi},
            {'lat': 1, 'lon': 1, 'timestamp': 1, 'speed': 1, 'course': 1, 'heading': 1, 'source_id': 1}
        ).sort('timestamp', -1).limit(limit).to_list(limit)
        
        # Remove duplicates (same timestamp from different sources) - keep first occurrence
        seen_timestamps = set()
        unique_positions = []
        for pos in positions:
            ts = pos.get('timestamp')
            if ts not in seen_timestamps:
                seen_timestamps.add(ts)
                unique_positions.append(pos)
        
        # Reverse to get chronological order (oldest to newest) for trail drawing
        unique_positions.reverse()
        
        serialized_positions = [serialize_doc(p) for p in unique_positions]
        
        logger.info(f"Track for {mmsi}: {len(unique_positions)} unique positions from {len(positions)} total")
        
        return {
            'mmsi': mmsi, 
            'track': serialized_positions,
            'count': len(serialized_positions),
            'total_records': len(positions)
        }
    except Exception as e:
        logger.error(f"Error loading track for {mmsi}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load track: {str(e)}")

@api_router.get("/history/{mmsi}")
async def get_vessel_history(mmsi: str):
    """Get complete historical data for an MMSI - properly merged from all sources"""
    try:
        # Get vessel info
        vessel = await db.vessels.find_one({'mmsi': mmsi})
        if not vessel:
            raise HTTPException(status_code=404, detail="Vessel not found")
        
        # Get all positions sorted chronologically (newest first for display)
        positions = await db.positions.find({'mmsi': mmsi}).sort('timestamp', -1).to_list(10000)
        
        # Get all messages sorted chronologically
        messages = await db.messages.find({'mmsi': mmsi}).sort('timestamp', -1).to_list(10000)
        
        # Count unique sources
        position_sources = set()
        for p in positions:
            if p.get('source_id'):
                position_sources.add(p['source_id'])
        
        message_sources = set()
        for m in messages:
            if m.get('source_id'):
                message_sources.add(m['source_id'])
        
        all_sources = position_sources.union(message_sources)
        
        # Serialize all data
        result = {
            'mmsi': mmsi,
            'vessel': serialize_doc(vessel),
            'positions': [serialize_doc(p) for p in positions],
            'messages': [serialize_doc(m) for m in messages],
            'position_count': len(positions),
            'message_count': len(messages),
            'unique_sources': len(all_sources),
            'source_ids': list(all_sources)
        }
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error loading history for {mmsi}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to load history: {str(e)}")

@api_router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket for real-time updates"""
    await manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            # Keep connection alive
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    global main_event_loop
    main_event_loop = asyncio.get_event_loop()
    logger.info("Main event loop captured for stream handlers")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
