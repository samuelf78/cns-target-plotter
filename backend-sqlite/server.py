"""
Complete SQLite Backend with MarineISA Integration
Full conversion from MongoDB backend - All functionality preserved
Runs on port 8002 alongside MongoDB backend on port 8001
"""

from fastapi import FastAPI, APIRouter, WebSocket, WebSocketDisconnect, UploadFile, File, HTTPException, BackgroundTasks, Response
from fastapi.responses import JSONResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
import aiosqlite
import os
from datetime import datetime, timezone, timedelta
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import asyncio
import json
from pyais import decode
from pyais.stream import TCPConnection, UDPReceiver, FileReaderStream
import socket
import serial
import serial.tools.list_ports
import threading
from openpyxl import Workbook
from io import BytesIO
import math

from marinesia_client import MarineISAClient

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Database setup
DATA_DIR = Path('/app/backend-sqlite/data')
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / 'ais_tracker.db'

# MarineISA setup
MARINESIA_ENABLED = os.getenv('MARINESIA_ENABLED', 'true').lower() == 'true'
MARINESIA_API_KEY = os.getenv('MARINESIA_API_KEY')
marinesia_client = None

if MARINESIA_ENABLED and MARINESIA_API_KEY:
    marinesia_client = MarineISAClient(MARINESIA_API_KEY)

# Create the main app
app = FastAPI(title="AIS SQLite Backend with MarineISA (Port 8002)")
api_router = APIRouter(prefix="/api")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# WebSocket connections manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except:
                pass

manager = ConnectionManager()

# Background stream controllers
active_streams = {}
main_event_loop = None
enrichment_queue = asyncio.Queue()

# Models
class VesselPosition(BaseModel):
    mmsi: str
    timestamp: datetime
    lat: float
    lon: float
    speed: Optional[float] = None
    course: Optional[float] = None
    heading: Optional[int] = None
    nav_status: Optional[int] = None

class VesselInfo(BaseModel):
    mmsi: str
    name: Optional[str] = None
    callsign: Optional[str] = None
    imo: Optional[int] = None
    ship_type: Optional[int] = None
    ship_type_text: Optional[str] = None
    dimension_a: Optional[int] = None
    dimension_b: Optional[int] = None
    dimension_c: Optional[int] = None
    dimension_d: Optional[int] = None
    last_seen: Optional[datetime] = None
    destination: Optional[str] = None
    eta: Optional[str] = None

class AISMessage(BaseModel):
    mmsi: str
    timestamp: datetime
    message_type: int
    raw: str
    decoded: Dict[str, Any]

class StreamConfig(BaseModel):
    stream_type: str  # tcp, udp, serial
    host: Optional[str] = None
    port: Optional[int] = None
    serial_port: Optional[str] = None
    baudrate: Optional[int] = 9600

class SearchQuery(BaseModel):
    mmsi: Optional[str] = None
    vessel_name: Optional[str] = None
    ship_type: Optional[int] = None
    start_time: Optional[datetime] = None
    end_time: Optional[datetime] = None

class DataSource(BaseModel):
    source_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source_type: str  # tcp, udp, serial, file
    name: str
    config: Dict[str, Any]
    status: str = "active"  # active, inactive, paused
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_message: Optional[datetime] = None
    message_count: int = 0
    target_count: int = 0  # Number of unique vessels
    fragment_count: int = 0  # Number of omitted/incomplete messages
    message_limit: int = 0  # Max messages to keep per source (0 = unlimited)
    target_limit: int = 0  # Max targets to display (0 = unlimited)
    keep_non_vessel_targets: bool = True  # Keep base stations and AtoNs regardless of target limit
    spoof_limit_km: float = 500.0  # Default 500km spoof limit
    is_paused: bool = False  # Pause state for streaming sources
    processing_complete: bool = False  # For file sources: true when fully processed

# Helper functions
def serialize_doc(doc):
    """Convert document to JSON-serializable dict"""
    if doc is None:
        return None
    
    if isinstance(doc, list):
        return [serialize_doc(item) for item in doc]
    
    if isinstance(doc, dict):
        serialized = {}
        for key, value in doc.items():
            if isinstance(value, dict):
                serialized[key] = serialize_doc(value)
            elif isinstance(value, list):
                serialized[key] = [serialize_doc(item) if isinstance(item, dict) else item for item in value]
            else:
                serialized[key] = value
        return serialized
    
    return doc

def sync_timestamp_with_message(decoded_msg):
    """Synchronize system timestamp with UTC second from AIS message"""
    now = datetime.now(timezone.utc)
    msg_type = decoded_msg.get('msg_type', 0)
    
    # Type 4, 11: Base Station Reports with full UTC timestamp
    if msg_type in [4, 11]:
        year = decoded_msg.get('year')
        month = decoded_msg.get('month')
        day = decoded_msg.get('day')
        hour = decoded_msg.get('hour')
        minute = decoded_msg.get('minute')
        second = decoded_msg.get('second')
        
        # Only use if all fields are valid
        if all(x is not None for x in [year, month, day, hour, minute, second]):
            try:
                # Validate ranges
                if (1970 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31 and
                    0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59):
                    return datetime(year, month, day, hour, minute, second, tzinfo=timezone.utc)
            except ValueError:
                pass
    
    # Types 1, 2, 3, 9, 18, 19, 21, 27: Have 'second' field only
    elif msg_type in [1, 2, 3, 9, 18, 19, 21, 27]:
        second = decoded_msg.get('second')
        if second is not None and 0 <= second <= 59:
            return now.replace(second=second, microsecond=0)
    
    return now

def get_mmsi_country(mmsi: str) -> str:
    """Get country from MMSI based on MID"""
    if not mmsi or len(mmsi) < 3:
        return "Unknown"
    
    mid = mmsi[:3]
    
    # MID to country mapping (abbreviated)
    mid_countries = {
        "201": "Albania", "202": "Andorra", "203": "Austria", "204": "Portugal", "205": "Belgium",
        "206": "Belarus", "207": "Bulgaria", "208": "Vatican", "209": "Cyprus", "210": "Cyprus",
        "211": "Germany", "212": "Cyprus", "213": "Georgia", "214": "Moldova", "215": "Malta",
        "216": "Armenia", "218": "Germany", "219": "Denmark", "220": "Denmark", "224": "Spain",
        "225": "Spain", "226": "France", "227": "France", "228": "France", "229": "Malta",
        "230": "Finland", "231": "Faroe Islands", "232": "United Kingdom", "233": "United Kingdom",
        "234": "United Kingdom", "235": "United Kingdom", "236": "Gibraltar", "237": "Greece",
        "238": "Croatia", "239": "Greece", "240": "Greece", "241": "Greece", "242": "Morocco",
        "243": "Hungary", "244": "Netherlands", "245": "Netherlands", "246": "Netherlands",
        "247": "Italy", "248": "Malta", "249": "Malta", "250": "Ireland", "251": "Iceland",
        "252": "Liechtenstein", "253": "Luxembourg", "254": "Monaco", "255": "Madeira",
        "256": "Malta", "257": "Norway", "258": "Norway", "259": "Norway", "261": "Poland",
        "262": "Montenegro", "263": "Portugal", "264": "Romania", "265": "Sweden", "266": "Sweden",
        "267": "Slovakia", "268": "San Marino", "269": "Switzerland", "270": "Czech Republic",
        "271": "Turkey", "272": "Ukraine", "273": "Russian Federation", "274": "North Macedonia",
        "275": "Latvia", "276": "Estonia", "277": "Lithuania", "278": "Slovenia", "279": "Serbia",
        "301": "Anguilla", "303": "Alaska", "304": "Antigua and Barbuda", "305": "Antigua and Barbuda",
        "306": "Netherlands Antilles", "307": "Aruba", "308": "Bahamas", "309": "Bahamas",
        "310": "Bermuda", "311": "Bahamas", "312": "Belize", "314": "Barbados", "316": "Canada",
        "319": "Cayman Islands", "321": "Costa Rica", "323": "Cuba", "325": "Dominica",
        "327": "Dominican Republic", "329": "Guadeloupe", "330": "Grenada", "331": "Greenland",
        "332": "Guatemala", "334": "Honduras", "336": "Haiti", "338": "United States",
        "339": "Jamaica", "341": "Saint Kitts and Nevis", "343": "Saint Lucia", "345": "Mexico",
        "347": "Martinique", "348": "Montserrat", "350": "Nicaragua", "351": "Panama",
        "352": "Panama", "353": "Panama", "354": "Panama", "355": "Panama", "356": "Panama",
        "357": "Panama", "358": "Puerto Rico", "359": "El Salvador", "361": "Saint Pierre and Miquelon",
        "362": "Trinidad and Tobago", "364": "Turks and Caicos Islands", "366": "United States",
        "367": "United States", "368": "United States", "369": "United States", "370": "Panama",
        "371": "Panama", "372": "Panama", "373": "Panama", "374": "Panama", "375": "Saint Vincent and the Grenadines",
        "376": "Saint Vincent and the Grenadines", "377": "Saint Vincent and the Grenadines",
        "378": "British Virgin Islands", "379": "United States Virgin Islands",
        "401": "Afghanistan", "403": "Saudi Arabia", "405": "Bangladesh", "408": "Bahrain",
        "410": "Bhutan", "412": "China", "413": "China", "414": "China", "416": "Taiwan",
        "417": "Sri Lanka", "419": "India", "422": "Iran", "423": "Azerbaijan", "425": "Iraq",
        "428": "Israel", "431": "Japan", "432": "Japan", "434": "Turkmenistan", "436": "Kazakhstan",
        "437": "Uzbekistan", "438": "Jordan", "440": "South Korea", "441": "South Korea",
        "443": "Palestine", "445": "North Korea", "447": "Kuwait", "450": "Lebanon",
        "451": "Kyrgyzstan", "453": "Macao", "455": "Maldives", "457": "Mongolia",
        "459": "Nepal", "461": "Oman", "463": "Pakistan", "466": "Qatar", "468": "Syria",
        "470": "United Arab Emirates", "472": "Tajikistan", "473": "Yemen", "475": "Yemen",
        "477": "Hong Kong", "478": "Bosnia and Herzegovina", "501": "Antarctica",
        "503": "Australia", "506": "Myanmar", "508": "Brunei", "510": "Micronesia",
        "511": "Palau", "512": "New Zealand", "514": "Cambodia", "515": "Cambodia",
        "516": "Christmas Island", "518": "Cook Islands", "520": "Fiji", "523": "Cocos Islands",
        "525": "Indonesia", "529": "Kiribati", "531": "Laos", "533": "Malaysia",
        "536": "Northern Mariana Islands", "538": "Marshall Islands", "540": "New Caledonia",
        "542": "Niue", "544": "Nauru", "546": "French Polynesia", "548": "Philippines",
        "553": "Papua New Guinea", "555": "Pitcairn Island", "557": "Solomon Islands",
        "559": "American Samoa", "561": "Samoa", "563": "Singapore", "564": "Singapore",
        "565": "Singapore", "566": "Singapore", "567": "Thailand", "570": "Tonga",
        "572": "Tuvalu", "574": "Vietnam", "576": "Vanuatu", "577": "Vanuatu", "578": "Wallis and Futuna",
        "601": "South Africa", "603": "Angola", "605": "Algeria", "607": "Saint Paul and Amsterdam Islands",
        "608": "Ascension Island", "609": "Burundi", "610": "Benin", "611": "Botswana",
        "612": "Central African Republic", "613": "Cameroon", "615": "Congo", "616": "Comoros",
        "617": "Cape Verde", "618": "Crozet Archipelago", "619": "Ivory Coast", "620": "Comoros",
        "621": "Djibouti", "622": "Egypt", "624": "Ethiopia", "625": "Eritrea", "626": "Gabonese Republic",
        "627": "Ghana", "629": "Gambia", "630": "Guinea-Bissau", "631": "Equatorial Guinea",
        "632": "Guinea", "633": "Burkina Faso", "634": "Kenya", "635": "Kerguelen Islands",
        "636": "Liberia", "637": "Liberia", "638": "South Sudan", "642": "Libya",
        "644": "Lesotho", "645": "Mauritius", "647": "Madagascar", "649": "Mali",
        "650": "Mozambique", "654": "Mauritania", "655": "Malawi", "656": "Niger",
        "657": "Nigeria", "659": "Namibia", "660": "Reunion", "661": "Rwanda",
        "662": "Sudan", "663": "Senegal", "664": "Seychelles", "665": "Saint Helena",
        "666": "Somalia", "667": "Sierra Leone", "668": "Sao Tome and Principe",
        "669": "Swaziland", "670": "Chad", "671": "Togolese Republic", "672": "Tunisia",
        "674": "Tanzania", "675": "Uganda", "676": "Democratic Republic of the Congo",
        "677": "Tanzania", "678": "Zambia", "679": "Zimbabwe"
    }
    
    return mid_countries.get(mid, "Unknown")

def get_ship_type_text(ship_type: int) -> str:
    """Convert ship type code to human readable text"""
    ship_types = {
        0: "Not available",
        20: "Wing in ground (WIG)",
        21: "Wing in ground (WIG), Hazardous category A",
        22: "Wing in ground (WIG), Hazardous category B", 
        23: "Wing in ground (WIG), Hazardous category C",
        24: "Wing in ground (WIG), Hazardous category D",
        30: "Fishing",
        31: "Towing",
        32: "Towing: length exceeds 200m or breadth exceeds 25m",
        33: "Dredging or underwater ops",
        34: "Diving ops",
        35: "Military ops",
        36: "Sailing",
        37: "Pleasure Craft",
        40: "High speed craft (HSC)",
        41: "High speed craft (HSC), Hazardous category A",
        42: "High speed craft (HSC), Hazardous category B",
        43: "High speed craft (HSC), Hazardous category C", 
        44: "High speed craft (HSC), Hazardous category D",
        50: "Pilot Vessel",
        51: "Search and Rescue vessel",
        52: "Tug",
        53: "Port Tender",
        54: "Anti-pollution equipment",
        55: "Law Enforcement",
        58: "Medical Transport",
        59: "Noncombatant ship according to RR Resolution No. 18",
        60: "Passenger",
        61: "Passenger, Hazardous category A",
        62: "Passenger, Hazardous category B",
        63: "Passenger, Hazardous category C",
        64: "Passenger, Hazardous category D",
        70: "Cargo",
        71: "Cargo, Hazardous category A",
        72: "Cargo, Hazardous category B",
        73: "Cargo, Hazardous category C",
        74: "Cargo, Hazardous category D",
        80: "Tanker",
        81: "Tanker, Hazardous category A",
        82: "Tanker, Hazardous category B",
        83: "Tanker, Hazardous category C",
        84: "Tanker, Hazardous category D",
        90: "Other Type",
        91: "Other Type, Hazardous category A",
        92: "Other Type, Hazardous category B",
        93: "Other Type, Hazardous category C",
        94: "Other Type, Hazardous category D"
    }
    
    return ship_types.get(ship_type, f"Unknown ({ship_type})")

def is_valid_position(lat: float, lon: float) -> bool:
    """Check if position is valid (not default/invalid values)"""
    # Check for invalid/default values
    if lat == 91.0 or lon == 181.0:  # AIS default invalid values
        return False
    if lat == 0.0 and lon == 0.0:  # Null island
        return False
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):  # Out of range
        return False
    return True

async def get_last_valid_position(mmsi: str) -> Optional[Dict[str, float]]:
    """Get the last valid position for a vessel"""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("""
            SELECT lat, lon FROM positions 
            WHERE mmsi = ? AND lat != 91.0 AND lon != 181.0 AND lat != 0.0 AND lon != 0.0
            ORDER BY timestamp DESC LIMIT 1
        """, (mmsi,)) as cursor:
            row = await cursor.fetchone()
            if row:
                return {"lat": row[0], "lon": row[1]}
    return None

async def backfill_invalid_positions(mmsi: str, valid_lat: float, valid_lon: float):
    """Backfill invalid positions with the last valid position"""
    async with aiosqlite.connect(DB_PATH) as db:
        # Find invalid positions for this MMSI
        async with db.execute("""
            SELECT id FROM positions 
            WHERE mmsi = ? AND (lat = 91.0 OR lon = 181.0 OR (lat = 0.0 AND lon = 0.0))
        """, (mmsi,)) as cursor:
            invalid_ids = [row[0] for row in await cursor.fetchall()]
        
        if invalid_ids:
            # Update invalid positions
            placeholders = ','.join('?' * len(invalid_ids))
            await db.execute(f"""
                UPDATE positions 
                SET lat = ?, lon = ? 
                WHERE id IN ({placeholders})
            """, (valid_lat, valid_lon, *invalid_ids))
            await db.commit()

def parse_log_line(raw_line: str):
    """Parse log line to extract timestamp and message"""
    line = raw_line.strip()
    
    # Try to extract timestamp from various log formats
    timestamp = None
    message = line
    
    # Format: "2024-01-15 10:30:45 !AIVDM,1,1,,A,15MwkT0P00G?Tq`K>P6B;wvP2<0=,0*23"
    if len(line) > 19 and line[4] == '-' and line[7] == '-' and line[10] == ' ':
        try:
            timestamp_str = line[:19]
            timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            message = line[20:].strip()
        except ValueError:
            pass
    
    return timestamp, message

async def process_ais_message(raw_message: str, source: str = "unknown", source_id: str = None, log_timestamp=None):
    """Process a single AIS message"""
    try:
        # Clean the message
        message = raw_message.strip()
        if not message:
            return
        
        # Skip non-AIS messages
        if not (message.startswith('!AIVDM') or message.startswith('!AIVDO')):
            return
        
        # Decode the AIS message
        try:
            decoded_msg = decode(message)
        except Exception as e:
            # Handle fragment or incomplete message
            if source_id:
                async with aiosqlite.connect(DB_PATH) as db:
                    await db.execute("""
                        UPDATE sources SET fragment_count = fragment_count + 1 
                        WHERE source_id = ?
                    """, (source_id,))
                    await db.commit()
            return
        
        # Get MMSI
        mmsi = str(decoded_msg.get('mmsi', ''))
        if not mmsi:
            return
        
        # Use log timestamp if provided, otherwise sync with message
        if log_timestamp:
            timestamp = log_timestamp
        else:
            timestamp = sync_timestamp_with_message(decoded_msg)
        
        # Store raw message
        async with aiosqlite.connect(DB_PATH) as db:
            message_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'message_type': decoded_msg.get('msg_type', 0),
                'raw': message,
                'decoded': json.dumps(decoded_msg),
                'source': source,
                'source_id': source_id or 'unknown'
            }
            
            await db.execute("""
                INSERT INTO messages (mmsi, timestamp, message_type, raw, decoded, source, source_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (message_doc['mmsi'], message_doc['timestamp'], message_doc['message_type'],
                  message_doc['raw'], message_doc['decoded'], message_doc['source'], message_doc['source_id']))
            
            # Update source stats
            if source_id:
                # Check message limit
                async with db.execute("""
                    SELECT message_limit FROM sources WHERE source_id = ?
                """, (source_id,)) as cursor:
                    row = await cursor.fetchone()
                    if row and row[0] > 0:
                        message_limit = row[0]
                        # Count current messages
                        async with db.execute("""
                            SELECT COUNT(*) FROM messages WHERE source_id = ?
                        """, (source_id,)) as cursor:
                            message_count = (await cursor.fetchone())[0]
                        
                        if message_count >= message_limit:
                            # Delete oldest messages
                            async with db.execute("""
                                SELECT id FROM messages WHERE source_id = ? 
                                ORDER BY timestamp ASC LIMIT ?
                            """, (source_id, message_count - message_limit + 1)) as cursor:
                                old_ids = [row[0] for row in await cursor.fetchall()]
                            
                            if old_ids:
                                placeholders = ','.join('?' * len(old_ids))
                                await db.execute(f"DELETE FROM messages WHERE id IN ({placeholders})", old_ids)
                                # Also delete related positions
                                await db.execute(f"DELETE FROM positions WHERE message_id IN ({placeholders})", old_ids)
                
                # Update source stats
                await db.execute("""
                    UPDATE sources 
                    SET last_message = ?, message_count = message_count + 1
                    WHERE source_id = ?
                """, (timestamp.isoformat(), source_id))
            
            await db.commit()
        
        # Process different message types
        msg_type = decoded_msg.get('msg_type', 0)
        
        # Position reports (Types 1, 2, 3)
        if msg_type in [1, 2, 3]:
            await process_position_report(decoded_msg, timestamp, source_id)
        
        # Base station report (Type 4)
        elif msg_type == 4:
            await process_base_station_report(decoded_msg, timestamp, source_id)
        
        # Static and voyage data (Type 5)
        elif msg_type == 5:
            await process_static_voyage_data(decoded_msg, timestamp, source_id)
        
        # Binary addressed message (Type 6)
        elif msg_type == 6:
            await process_binary_message(decoded_msg, timestamp, source_id)
        
        # Binary acknowledge (Type 7)
        elif msg_type == 7:
            await process_binary_ack(decoded_msg, timestamp, source_id)
        
        # Binary broadcast message (Type 8)
        elif msg_type == 8:
            await process_binary_broadcast(decoded_msg, timestamp, source_id)
        
        # Standard SAR aircraft position report (Type 9)
        elif msg_type == 9:
            await process_sar_aircraft_report(decoded_msg, timestamp, source_id)
        
        # UTC and date inquiry (Type 10)
        elif msg_type == 10:
            await process_utc_inquiry(decoded_msg, timestamp, source_id)
        
        # UTC and date response (Type 11)
        elif msg_type == 11:
            await process_utc_response(decoded_msg, timestamp, source_id)
        
        # Addressed safety related message (Type 12)
        elif msg_type == 12:
            await process_safety_message(decoded_msg, timestamp, source_id)
        
        # Safety related acknowledgement (Type 13)
        elif msg_type == 13:
            await process_safety_ack(decoded_msg, timestamp, source_id)
        
        # Safety related broadcast message (Type 14)
        elif msg_type == 14:
            await process_safety_broadcast(decoded_msg, timestamp, source_id)
        
        # Interrogation (Type 15)
        elif msg_type == 15:
            await process_interrogation(decoded_msg, timestamp, source_id)
        
        # Assignment mode command (Type 16)
        elif msg_type == 16:
            await process_assignment_command(decoded_msg, timestamp, source_id)
        
        # DGNSS binary broadcast message (Type 17)
        elif msg_type == 17:
            await process_dgnss_broadcast(decoded_msg, timestamp, source_id)
        
        # Standard Class B CS position report (Type 18)
        elif msg_type == 18:
            await process_class_b_position(decoded_msg, timestamp, source_id)
        
        # Extended Class B equipment position report (Type 19)
        elif msg_type == 19:
            await process_extended_class_b_position(decoded_msg, timestamp, source_id)
        
        # Data link management (Type 20)
        elif msg_type == 20:
            await process_data_link_management(decoded_msg, timestamp, source_id)
        
        # Aid-to-navigation report (Type 21)
        elif msg_type == 21:
            await process_aid_to_navigation(decoded_msg, timestamp, source_id)
        
        # Channel management (Type 22)
        elif msg_type == 22:
            await process_channel_management(decoded_msg, timestamp, source_id)
        
        # Group assignment command (Type 23)
        elif msg_type == 23:
            await process_group_assignment(decoded_msg, timestamp, source_id)
        
        # Static data report (Type 24)
        elif msg_type == 24:
            await process_static_data_report(decoded_msg, timestamp, source_id)
        
        # Single slot binary message (Type 25)
        elif msg_type == 25:
            await process_single_slot_binary(decoded_msg, timestamp, source_id)
        
        # Multiple slot binary message (Type 26)
        elif msg_type == 26:
            await process_multiple_slot_binary(decoded_msg, timestamp, source_id)
        
        # Position report for long-range applications (Type 27)
        elif msg_type == 27:
            await process_long_range_position(decoded_msg, timestamp, source_id)
        
    except Exception as e:
        logging.error(f"Error processing AIS message: {e}")

async def process_position_report(decoded_msg, timestamp, source_id):
    """Process position report messages (Types 1, 2, 3)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is None or lon is None:
        return
    
    # Check for spoof detection
    if source_id:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute("""
                SELECT spoof_limit_km FROM sources WHERE source_id = ?
            """, (source_id,)) as cursor:
                row = await cursor.fetchone()
                if row:
                    spoof_limit_km = row[0]
                    
                    # Get last position
                    last_pos = await get_last_valid_position(mmsi)
                    if last_pos and is_valid_position(lat, lon):
                        # Calculate distance
                        distance = calculate_distance(
                            last_pos['lat'], last_pos['lon'], lat, lon
                        )
                        if distance > spoof_limit_km:
                            # Mark as potential spoof
                            logging.warning(f"Potential spoof detected for {mmsi}: {distance:.1f}km jump")
    
    # Store position
    async with aiosqlite.connect(DB_PATH) as db:
        position_doc = {
            'mmsi': mmsi,
            'timestamp': timestamp.isoformat(),
            'lat': lat,
            'lon': lon,
            'speed': decoded_msg.get('speed'),
            'course': decoded_msg.get('course'),
            'heading': decoded_msg.get('heading'),
            'nav_status': decoded_msg.get('nav_status'),
            'source_id': source_id
        }
        
        await db.execute("""
            INSERT INTO positions (mmsi, timestamp, lat, lon, speed, course, heading, nav_status, source_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'], 
              position_doc['lon'], position_doc['speed'], position_doc['course'],
              position_doc['heading'], position_doc['nav_status'], position_doc['source_id']))
        
        # Update vessel info
        if is_valid_position(lat, lon):
            # Count positions for this vessel
            async with db.execute("""
                SELECT COUNT(*) FROM positions WHERE mmsi = ?
            """, (mmsi,)) as cursor:
                pos_count = (await cursor.fetchone())[0]
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, last_seen, lat, lon, speed, course, heading, nav_status, position_count, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, timestamp.isoformat(), lat, lon, decoded_msg.get('speed'),
                  decoded_msg.get('course'), decoded_msg.get('heading'), 
                  decoded_msg.get('nav_status'), pos_count, get_mmsi_country(mmsi)))
        else:
            # Invalid position - try to backfill
            last_valid = await get_last_valid_position(mmsi)
            if last_valid:
                await backfill_invalid_positions(mmsi, last_valid['lat'], last_valid['lon'])
        
        await db.commit()
        
        # Broadcast to WebSocket clients
        await manager.broadcast({
            'type': 'position_update',
            'data': serialize_doc(position_doc)
        })
        
        # Queue for enrichment
        if marinesia_client:
            try:
                enrichment_queue.put_nowait(mmsi)
            except:
                pass

async def process_base_station_report(decoded_msg, timestamp, source_id):
    """Process base station report (Type 4)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, source_id)
                VALUES (?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], 
                  position_doc['lat'], position_doc['lon'], position_doc['source_id']))
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, last_seen, lat, lon, country, ship_type, ship_type_text)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, timestamp.isoformat(), lat, lon, get_mmsi_country(mmsi), 0, "Base Station"))
            
            await db.commit()
            
            # Broadcast to WebSocket clients
            await manager.broadcast({
                'type': 'base_station_update',
                'data': serialize_doc(position_doc)
            })

async def process_static_voyage_data(decoded_msg, timestamp, source_id):
    """Process static and voyage data (Type 5)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    
    async with aiosqlite.connect(DB_PATH) as db:
        vessel_data = {
            'mmsi': mmsi,
            'name': decoded_msg.get('shipname', '').strip(),
            'callsign': decoded_msg.get('callsign', '').strip(),
            'imo': decoded_msg.get('imo'),
            'ship_type': decoded_msg.get('ship_type'),
            'ship_type_text': get_ship_type_text(decoded_msg.get('ship_type', 0)),
            'dimension_a': decoded_msg.get('to_bow'),
            'dimension_b': decoded_msg.get('to_stern'),
            'dimension_c': decoded_msg.get('to_port'),
            'dimension_d': decoded_msg.get('to_starboard'),
            'destination': decoded_msg.get('destination', '').strip(),
            'eta': decoded_msg.get('eta'),
            'country': get_mmsi_country(mmsi),
            'last_seen': timestamp.isoformat()
        }
        
        # Update vessel with static data
        await db.execute("""
            INSERT OR REPLACE INTO vessels 
            (mmsi, name, callsign, imo, ship_type, ship_type_text, dimension_a, dimension_b, 
             dimension_c, dimension_d, destination, eta, country, last_seen)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (vessel_data['mmsi'], vessel_data['name'], vessel_data['callsign'],
              vessel_data['imo'], vessel_data['ship_type'], vessel_data['ship_type_text'],
              vessel_data['dimension_a'], vessel_data['dimension_b'], vessel_data['dimension_c'],
              vessel_data['dimension_d'], vessel_data['destination'], vessel_data['eta'],
              vessel_data['country'], vessel_data['last_seen']))
        
        await db.commit()
        
        # Queue for enrichment
        if marinesia_client:
            try:
                enrichment_queue.put_nowait(mmsi)
            except:
                pass

# Placeholder functions for other message types
async def process_binary_message(decoded_msg, timestamp, source_id):
    """Process binary addressed message (Type 6)"""
    pass

async def process_binary_ack(decoded_msg, timestamp, source_id):
    """Process binary acknowledge (Type 7)"""
    pass

async def process_binary_broadcast(decoded_msg, timestamp, source_id):
    """Process binary broadcast message (Type 8)"""
    pass

async def process_sar_aircraft_report(decoded_msg, timestamp, source_id):
    """Process SAR aircraft position report (Type 9)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'speed': decoded_msg.get('speed'),
                'course': decoded_msg.get('course'),
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, speed, course, source_id)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'],
                  position_doc['lon'], position_doc['speed'], position_doc['course'], position_doc['source_id']))
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, last_seen, lat, lon, speed, course, country, ship_type, ship_type_text)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, timestamp.isoformat(), lat, lon, decoded_msg.get('speed'),
                  decoded_msg.get('course'), get_mmsi_country(mmsi), 9, "SAR Aircraft"))
            
            await db.commit()
            
            # Broadcast to WebSocket clients
            await manager.broadcast({
                'type': 'sar_aircraft_update',
                'data': serialize_doc(position_doc)
            })

async def process_utc_inquiry(decoded_msg, timestamp, source_id):
    """Process UTC and date inquiry (Type 10)"""
    pass

async def process_utc_response(decoded_msg, timestamp, source_id):
    """Process UTC and date response (Type 11)"""
    pass

async def process_safety_message(decoded_msg, timestamp, source_id):
    """Process addressed safety related message (Type 12)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    text = decoded_msg.get('text', '').strip()
    
    if text:
        async with aiosqlite.connect(DB_PATH) as db:
            text_msg_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'message_type': 12,
                'text': text,
                'dest_mmsi': decoded_msg.get('dest_mmsi'),
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO text_messages (mmsi, timestamp, message_type, text, dest_mmsi, source_id)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (text_msg_doc['mmsi'], text_msg_doc['timestamp'], text_msg_doc['message_type'],
                  text_msg_doc['text'], text_msg_doc['dest_mmsi'], text_msg_doc['source_id']))
            
            await db.commit()

async def process_safety_ack(decoded_msg, timestamp, source_id):
    """Process safety related acknowledgement (Type 13)"""
    pass

async def process_safety_broadcast(decoded_msg, timestamp, source_id):
    """Process safety related broadcast message (Type 14)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    text = decoded_msg.get('text', '').strip()
    
    if text:
        async with aiosqlite.connect(DB_PATH) as db:
            text_msg_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'message_type': 14,
                'text': text,
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO text_messages (mmsi, timestamp, message_type, text, source_id)
                VALUES (?, ?, ?, ?, ?)
            """, (text_msg_doc['mmsi'], text_msg_doc['timestamp'], text_msg_doc['message_type'],
                  text_msg_doc['text'], text_msg_doc['source_id']))
            
            await db.commit()

async def process_interrogation(decoded_msg, timestamp, source_id):
    """Process interrogation (Type 15)"""
    pass

async def process_assignment_command(decoded_msg, timestamp, source_id):
    """Process assignment mode command (Type 16)"""
    pass

async def process_dgnss_broadcast(decoded_msg, timestamp, source_id):
    """Process DGNSS binary broadcast message (Type 17)"""
    pass

async def process_class_b_position(decoded_msg, timestamp, source_id):
    """Process Standard Class B CS position report (Type 18)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'speed': decoded_msg.get('speed'),
                'course': decoded_msg.get('course'),
                'heading': decoded_msg.get('heading'),
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, speed, course, heading, source_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'],
                  position_doc['lon'], position_doc['speed'], position_doc['course'],
                  position_doc['heading'], position_doc['source_id']))
            
            # Count positions for this vessel
            async with db.execute("""
                SELECT COUNT(*) FROM positions WHERE mmsi = ?
            """, (mmsi,)) as cursor:
                pos_count = (await cursor.fetchone())[0]
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, last_seen, lat, lon, speed, course, heading, position_count, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, timestamp.isoformat(), lat, lon, decoded_msg.get('speed'),
                  decoded_msg.get('course'), decoded_msg.get('heading'), pos_count, get_mmsi_country(mmsi)))
            
            await db.commit()
            
            # Broadcast to WebSocket clients
            await manager.broadcast({
                'type': 'class_b_position_update',
                'data': serialize_doc(position_doc)
            })
            
            # Queue for enrichment
            if marinesia_client:
                try:
                    enrichment_queue.put_nowait(mmsi)
                except:
                    pass

async def process_extended_class_b_position(decoded_msg, timestamp, source_id):
    """Process Extended Class B equipment position report (Type 19)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'speed': decoded_msg.get('speed'),
                'course': decoded_msg.get('course'),
                'heading': decoded_msg.get('heading'),
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, speed, course, heading, source_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'],
                  position_doc['lon'], position_doc['speed'], position_doc['course'],
                  position_doc['heading'], position_doc['source_id']))
            
            # Update vessel with extended info
            vessel_data = {
                'mmsi': mmsi,
                'name': decoded_msg.get('shipname', '').strip(),
                'ship_type': decoded_msg.get('ship_type'),
                'ship_type_text': get_ship_type_text(decoded_msg.get('ship_type', 0)),
                'dimension_a': decoded_msg.get('to_bow'),
                'dimension_b': decoded_msg.get('to_stern'),
                'dimension_c': decoded_msg.get('to_port'),
                'dimension_d': decoded_msg.get('to_starboard'),
                'country': get_mmsi_country(mmsi),
                'last_seen': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'speed': decoded_msg.get('speed'),
                'course': decoded_msg.get('course'),
                'heading': decoded_msg.get('heading')
            }
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, name, ship_type, ship_type_text, dimension_a, dimension_b, dimension_c, dimension_d,
                 country, last_seen, lat, lon, speed, course, heading)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (vessel_data['mmsi'], vessel_data['name'], vessel_data['ship_type'],
                  vessel_data['ship_type_text'], vessel_data['dimension_a'], vessel_data['dimension_b'],
                  vessel_data['dimension_c'], vessel_data['dimension_d'], vessel_data['country'],
                  vessel_data['last_seen'], vessel_data['lat'], vessel_data['lon'],
                  vessel_data['speed'], vessel_data['course'], vessel_data['heading']))
            
            await db.commit()
            
            # Queue for enrichment
            if marinesia_client:
                try:
                    enrichment_queue.put_nowait(mmsi)
                except:
                    pass

async def process_data_link_management(decoded_msg, timestamp, source_id):
    """Process data link management (Type 20)"""
    pass

async def process_aid_to_navigation(decoded_msg, timestamp, source_id):
    """Process aid-to-navigation report (Type 21)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, source_id)
                VALUES (?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'],
                  position_doc['lon'], position_doc['source_id']))
            
            # Update vessel as AtoN
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, name, last_seen, lat, lon, country, ship_type, ship_type_text)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, decoded_msg.get('name', '').strip(), timestamp.isoformat(), lat, lon,
                  get_mmsi_country(mmsi), 21, "Aid to Navigation"))
            
            await db.commit()

async def process_channel_management(decoded_msg, timestamp, source_id):
    """Process channel management (Type 22)"""
    pass

async def process_group_assignment(decoded_msg, timestamp, source_id):
    """Process group assignment command (Type 23)"""
    pass

async def process_static_data_report(decoded_msg, timestamp, source_id):
    """Process static data report (Type 24)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    part_num = decoded_msg.get('part_num', 0)
    
    async with aiosqlite.connect(DB_PATH) as db:
        if part_num == 0:  # Part A - vessel name
            vessel_name = decoded_msg.get('shipname', '').strip()
            if vessel_name:
                await db.execute("""
                    INSERT OR REPLACE INTO vessels (mmsi, name, country, last_seen)
                    VALUES (?, ?, ?, ?)
                """, (mmsi, vessel_name, get_mmsi_country(mmsi), timestamp.isoformat()))
        
        elif part_num == 1:  # Part B - vessel dimensions and type
            vessel_data = {
                'mmsi': mmsi,
                'ship_type': decoded_msg.get('ship_type'),
                'ship_type_text': get_ship_type_text(decoded_msg.get('ship_type', 0)),
                'dimension_a': decoded_msg.get('to_bow'),
                'dimension_b': decoded_msg.get('to_stern'),
                'dimension_c': decoded_msg.get('to_port'),
                'dimension_d': decoded_msg.get('to_starboard'),
                'callsign': decoded_msg.get('callsign', '').strip(),
                'country': get_mmsi_country(mmsi),
                'last_seen': timestamp.isoformat()
            }
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, ship_type, ship_type_text, dimension_a, dimension_b, dimension_c, dimension_d,
                 callsign, country, last_seen)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (vessel_data['mmsi'], vessel_data['ship_type'], vessel_data['ship_type_text'],
                  vessel_data['dimension_a'], vessel_data['dimension_b'], vessel_data['dimension_c'],
                  vessel_data['dimension_d'], vessel_data['callsign'], vessel_data['country'],
                  vessel_data['last_seen']))
        
        await db.commit()
        
        # Queue for enrichment
        if marinesia_client:
            try:
                enrichment_queue.put_nowait(mmsi)
            except:
                pass

async def process_single_slot_binary(decoded_msg, timestamp, source_id):
    """Process single slot binary message (Type 25)"""
    pass

async def process_multiple_slot_binary(decoded_msg, timestamp, source_id):
    """Process multiple slot binary message (Type 26)"""
    pass

async def process_long_range_position(decoded_msg, timestamp, source_id):
    """Process position report for long-range applications (Type 27)"""
    mmsi = str(decoded_msg.get('mmsi', ''))
    lat = decoded_msg.get('lat')
    lon = decoded_msg.get('lon')
    
    if lat is not None and lon is not None and is_valid_position(lat, lon):
        async with aiosqlite.connect(DB_PATH) as db:
            position_doc = {
                'mmsi': mmsi,
                'timestamp': timestamp.isoformat(),
                'lat': lat,
                'lon': lon,
                'speed': decoded_msg.get('speed'),
                'course': decoded_msg.get('course'),
                'nav_status': decoded_msg.get('nav_status'),
                'source_id': source_id
            }
            
            await db.execute("""
                INSERT INTO positions (mmsi, timestamp, lat, lon, speed, course, nav_status, source_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (position_doc['mmsi'], position_doc['timestamp'], position_doc['lat'],
                  position_doc['lon'], position_doc['speed'], position_doc['course'],
                  position_doc['nav_status'], position_doc['source_id']))
            
            await db.execute("""
                INSERT OR REPLACE INTO vessels 
                (mmsi, last_seen, lat, lon, speed, course, nav_status, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (mmsi, timestamp.isoformat(), lat, lon, decoded_msg.get('speed'),
                  decoded_msg.get('course'), decoded_msg.get('nav_status'), get_mmsi_country(mmsi)))
            
            await db.commit()

def calculate_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two points in kilometers using Haversine formula"""
    R = 6371  # Earth's radius in kilometers
    
    lat1_rad = math.radians(lat1)
    lon1_rad = math.radians(lon1)
    lat2_rad = math.radians(lat2)
    lon2_rad = math.radians(lon2)
    
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    
    a = math.sin(dlat/2)**2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    
    return R * c

async def enrichment_worker():
    """Background worker to enrich vessels with MarineISA data"""
    while True:
        try:
            mmsi = await enrichment_queue.get()
            
            if not marinesia_client:
                continue
            
            # Check if already enriched recently
            async with aiosqlite.connect(DB_PATH) as db:
                async with db.execute("""
                    SELECT enriched_at FROM vessel_enrichment WHERE mmsi = ?
                """, (mmsi,)) as cursor:
                    row = await cursor.fetchone()
                    if row:
                        # Already enriched, skip
                        continue
            
            # Enrich vessel
            enriched_data = await marinesia_client.enrich_vessel(mmsi)
            
            if enriched_data.get('enriched'):
                # Store enrichment
                async with aiosqlite.connect(DB_PATH) as db:
                    await db.execute("""
                        INSERT OR REPLACE INTO vessel_enrichment 
                        (mmsi, profile_data, image_url, enriched_at)
                        VALUES (?, ?, ?, ?)
                    """, (
                        mmsi,
                        json.dumps(enriched_data.get('profile')),
                        enriched_data.get('image_url'),
                        enriched_data.get('enriched_at')
                    ))
                    await db.commit()
                
        except Exception as e:
            logging.error(f"Enrichment error: {e}")
        
        await asyncio.sleep(0.1)

async def init_db():
    """Initialize SQLite database with all required tables"""
    async with aiosqlite.connect(DB_PATH) as db:
        # Vessels table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS vessels (
                mmsi TEXT PRIMARY KEY,
                name TEXT,
                callsign TEXT,
                imo INTEGER,
                ship_type INTEGER,
                ship_type_text TEXT,
                dimension_a INTEGER,
                dimension_b INTEGER,
                dimension_c INTEGER,
                dimension_d INTEGER,
                destination TEXT,
                eta TEXT,
                country TEXT,
                last_seen TEXT,
                lat REAL,
                lon REAL,
                speed REAL,
                course REAL,
                heading INTEGER,
                nav_status INTEGER,
                position_count INTEGER DEFAULT 0
            )
        """)
        
        # Positions table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS positions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mmsi TEXT,
                timestamp TEXT,
                lat REAL,
                lon REAL,
                speed REAL,
                course REAL,
                heading INTEGER,
                nav_status INTEGER,
                source_id TEXT,
                message_id INTEGER,
                FOREIGN KEY (mmsi) REFERENCES vessels (mmsi)
            )
        """)
        
        # Messages table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mmsi TEXT,
                timestamp TEXT,
                message_type INTEGER,
                raw TEXT,
                decoded TEXT,
                source TEXT,
                source_id TEXT,
                FOREIGN KEY (mmsi) REFERENCES vessels (mmsi)
            )
        """)
        
        # Text messages table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS text_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                mmsi TEXT,
                timestamp TEXT,
                message_type INTEGER,
                text TEXT,
                dest_mmsi TEXT,
                source_id TEXT,
                FOREIGN KEY (mmsi) REFERENCES vessels (mmsi)
            )
        """)
        
        # Sources table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS sources (
                source_id TEXT PRIMARY KEY,
                source_type TEXT,
                name TEXT,
                config TEXT,
                status TEXT DEFAULT 'active',
                created_at TEXT,
                last_message TEXT,
                message_count INTEGER DEFAULT 0,
                target_count INTEGER DEFAULT 0,
                fragment_count INTEGER DEFAULT 0,
                message_limit INTEGER DEFAULT 0,
                target_limit INTEGER DEFAULT 0,
                keep_non_vessel_targets BOOLEAN DEFAULT TRUE,
                spoof_limit_km REAL DEFAULT 500.0,
                is_paused BOOLEAN DEFAULT FALSE,
                processing_complete BOOLEAN DEFAULT FALSE
            )
        """)
        
        # Vessel enrichment table
        await db.execute("""
            CREATE TABLE IF NOT EXISTS vessel_enrichment (
                mmsi TEXT PRIMARY KEY,
                profile_data TEXT,
                image_url TEXT,
                enriched_at TEXT,
                FOREIGN KEY (mmsi) REFERENCES vessels (mmsi)
            )
        """)
        
        # Create indexes for performance
        await db.execute("CREATE INDEX IF NOT EXISTS idx_positions_mmsi ON positions (mmsi)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_positions_timestamp ON positions (timestamp)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_messages_mmsi ON messages (mmsi)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_messages_timestamp ON messages (timestamp)")
        await db.execute("CREATE INDEX IF NOT EXISTS idx_vessels_last_seen ON vessels (last_seen)")
        
        await db.commit()

# Streaming functions
async def tcp_stream_worker(host: str, port: int, source_id: str):
    """Worker for TCP AIS stream"""
    try:
        reader, writer = await asyncio.open_connection(host, port)
        
        while source_id in active_streams and active_streams[source_id].get('running', False):
            try:
                # Check if paused
                async with aiosqlite.connect(DB_PATH) as db:
                    async with db.execute("""
                        SELECT is_paused FROM sources WHERE source_id = ?
                    """, (source_id,)) as cursor:
                        row = await cursor.fetchone()
                        if row and row[0]:
                            await asyncio.sleep(1)
                            continue
                
                data = await reader.readline()
                if not data:
                    break
                
                message = data.decode('utf-8', errors='ignore').strip()
                if message:
                    await process_ais_message(message, f"tcp://{host}:{port}", source_id)
                
            except Exception as e:
                logging.error(f"TCP stream error: {e}")
                break
        
        writer.close()
        await writer.wait_closed()
        
    except Exception as e:
        logging.error(f"TCP connection error: {e}")
    finally:
        if source_id in active_streams:
            active_streams[source_id]['running'] = False

async def udp_stream_worker(host: str, port: int, source_id: str):
    """Worker for UDP AIS stream"""
    try:
        # Create UDP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.bind((host, port))
        sock.setblocking(False)
        
        while source_id in active_streams and active_streams[source_id].get('running', False):
            try:
                # Check if paused
                async with aiosqlite.connect(DB_PATH) as db:
                    async with db.execute("""
                        SELECT is_paused FROM sources WHERE source_id = ?
                    """, (source_id,)) as cursor:
                        row = await cursor.fetchone()
                        if row and row[0]:
                            await asyncio.sleep(1)
                            continue
                
                data, addr = await asyncio.get_event_loop().sock_recvfrom(sock, 1024)
                message = data.decode('utf-8', errors='ignore').strip()
                if message:
                    await process_ais_message(message, f"udp://{host}:{port}", source_id)
                
            except Exception as e:
                await asyncio.sleep(0.1)
        
        sock.close()
        
    except Exception as e:
        logging.error(f"UDP stream error: {e}")
    finally:
        if source_id in active_streams:
            active_streams[source_id]['running'] = False

async def serial_stream_worker(port: str, baudrate: int, source_id: str):
    """Worker for Serial AIS stream"""
    try:
        ser = serial.Serial(port, baudrate, timeout=1)
        
        while source_id in active_streams and active_streams[source_id].get('running', False):
            try:
                # Check if paused
                async with aiosqlite.connect(DB_PATH) as db:
                    async with db.execute("""
                        SELECT is_paused FROM sources WHERE source_id = ?
                    """, (source_id,)) as cursor:
                        row = await cursor.fetchone()
                        if row and row[0]:
                            await asyncio.sleep(1)
                            continue
                
                line = ser.readline()
                if line:
                    message = line.decode('utf-8', errors='ignore').strip()
                    if message:
                        await process_ais_message(message, f"serial://{port}", source_id)
                
            except Exception as e:
                logging.error(f"Serial stream error: {e}")
                break
        
        ser.close()
        
    except Exception as e:
        logging.error(f"Serial connection error: {e}")
    finally:
        if source_id in active_streams:
            active_streams[source_id]['running'] = False

async def file_stream_worker(file_path: str, source_id: str):
    """Worker for file-based AIS stream"""
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            line_count = 0
            for line in f:
                if source_id not in active_streams or not active_streams[source_id].get('running', False):
                    break
                
                # Check if paused
                async with aiosqlite.connect(DB_PATH) as db:
                    async with db.execute("""
                        SELECT is_paused FROM sources WHERE source_id = ?
                    """, (source_id,)) as cursor:
                        row = await cursor.fetchone()
                        if row and row[0]:
                            await asyncio.sleep(1)
                            continue
                
                line = line.strip()
                if line:
                    # Parse timestamp from log line if present
                    log_timestamp, message = parse_log_line(line)
                    await process_ais_message(message, f"file://{file_path}", source_id, log_timestamp)
                
                line_count += 1
                if line_count % 100 == 0:  # Small delay every 100 lines
                    await asyncio.sleep(0.01)
        
        # Mark as complete
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("""
                UPDATE sources SET processing_complete = TRUE WHERE source_id = ?
            """, (source_id,))
            await db.commit()
        
    except Exception as e:
        logging.error(f"File stream error: {e}")
    finally:
        if source_id in active_streams:
            active_streams[source_id]['running'] = False

# API Endpoints
@app.on_event("startup")
async def startup():
    global main_event_loop
    main_event_loop = asyncio.get_event_loop()
    await init_db()
    # Start enrichment worker
    asyncio.create_task(enrichment_worker())

@api_router.get("/")
async def root():
    return {
        "message": "AIS SQLite Backend with MarineISA Integration",
        "port": 8002,
        "database": str(DB_PATH),
        "marinesia_enabled": MARINESIA_ENABLED
    }

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...), background_tasks: BackgroundTasks = None):
    """Upload and process AIS data file"""
    if not file.filename.endswith(('.txt', '.log', '.ais')):
        raise HTTPException(status_code=400, detail="Invalid file type. Only .txt, .log, and .ais files are supported.")
    
    # Save uploaded file
    upload_dir = Path("/tmp/ais_uploads")
    upload_dir.mkdir(exist_ok=True)
    file_path = upload_dir / f"{uuid.uuid4()}_{file.filename}"
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    # Create source record
    source_id = str(uuid.uuid4())
    source_doc = DataSource(
        source_id=source_id,
        source_type="file",
        name=file.filename,
        config={"file_path": str(file_path), "file_size": len(content)},
        status="active"
    )
    
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            INSERT INTO sources 
            (source_id, source_type, name, config, status, created_at, message_count, target_count, fragment_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (source_doc.source_id, source_doc.source_type, source_doc.name, 
              json.dumps(source_doc.config), source_doc.status, source_doc.created_at.isoformat(),
              source_doc.message_count, source_doc.target_count, source_doc.fragment_count))
        await db.commit()
    
    # Start processing in background
    active_streams[source_id] = {'running': True, 'type': 'file'}
    if background_tasks:
        background_tasks.add_task(file_stream_worker, str(file_path), source_id)
    else:
        asyncio.create_task(file_stream_worker(str(file_path), source_id))
    
    return {
        "message": f"File {file.filename} uploaded successfully",
        "source_id": source_id,
        "file_size": len(content)
    }

@api_router.post("/stream/start")
async def start_stream(config: StreamConfig, background_tasks: BackgroundTasks):
    """Start a new AIS data stream"""
    source_id = str(uuid.uuid4())
    
    # Create source record
    source_doc = DataSource(
        source_id=source_id,
        source_type=config.stream_type,
        name=f"{config.stream_type.upper()} Stream",
        config=config.dict(),
        status="active"
    )
    
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            INSERT INTO sources 
            (source_id, source_type, name, config, status, created_at, message_count, target_count, fragment_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (source_doc.source_id, source_doc.source_type, source_doc.name,
              json.dumps(source_doc.config), source_doc.status, source_doc.created_at.isoformat(),
              source_doc.message_count, source_doc.target_count, source_doc.fragment_count))
        await db.commit()
    
    # Start appropriate stream worker
    active_streams[source_id] = {'running': True, 'type': config.stream_type}
    
    if config.stream_type == "tcp":
        background_tasks.add_task(tcp_stream_worker, config.host, config.port, source_id)
    elif config.stream_type == "udp":
        background_tasks.add_task(udp_stream_worker, config.host, config.port, source_id)
    elif config.stream_type == "serial":
        background_tasks.add_task(serial_stream_worker, config.serial_port, config.baudrate, source_id)
    else:
        raise HTTPException(status_code=400, detail="Invalid stream type")
    
    return {
        "message": f"{config.stream_type.upper()} stream started",
        "source_id": source_id,
        "config": config.dict()
    }

@api_router.post("/stream/stop/{stream_id}")
async def stop_stream(stream_id: str):
    """Stop an active stream"""
    if stream_id in active_streams:
        active_streams[stream_id]['running'] = False
        del active_streams[stream_id]
        return {"message": f"Stream {stream_id} stopped"}
    else:
        raise HTTPException(status_code=404, detail="Stream not found")

@api_router.get("/stream/active")
async def get_active_streams():
    """Get list of active streams"""
    return {"active_streams": list(active_streams.keys())}

@api_router.get("/sources")
async def get_sources():
    """Get all data sources"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM sources ORDER BY created_at DESC") as cursor:
            sources = []
            async for row in cursor:
                source = dict(row)
                source['config'] = json.loads(source['config'])
                sources.append(source)
            return {"sources": sources}

@api_router.patch("/sources/{source_id}/toggle")
async def toggle_source(source_id: str):
    """Toggle source active/inactive status"""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT status FROM sources WHERE source_id = ?", (source_id,)) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Source not found")
            
            new_status = "inactive" if row[0] == "active" else "active"
            await db.execute("UPDATE sources SET status = ? WHERE source_id = ?", (new_status, source_id))
            await db.commit()
            
            return {"message": f"Source {source_id} is now {new_status}"}

@api_router.patch("/sources/{source_id}/spoof-limit")
async def update_spoof_limit(source_id: str, spoof_limit_km: float):
    """Update spoof detection limit for a source"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET spoof_limit_km = ? WHERE source_id = ?", (spoof_limit_km, source_id))
        await db.commit()
        return {"message": f"Spoof limit updated to {spoof_limit_km}km"}

@api_router.post("/sources/{source_id}/pause")
async def pause_source(source_id: str):
    """Pause a streaming source"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET is_paused = TRUE WHERE source_id = ?", (source_id,))
        await db.commit()
        return {"message": f"Source {source_id} paused"}

@api_router.post("/sources/{source_id}/resume")
async def resume_source(source_id: str):
    """Resume a paused source"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET is_paused = FALSE WHERE source_id = ?", (source_id,))
        await db.commit()
        return {"message": f"Source {source_id} resumed"}

@api_router.patch("/sources/{source_id}/message-limit")
async def update_message_limit(source_id: str, message_limit: int):
    """Update message limit for a source"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET message_limit = ? WHERE source_id = ?", (message_limit, source_id))
        await db.commit()
        return {"message": f"Message limit updated to {message_limit}"}

@api_router.patch("/sources/{source_id}/target-limit")
async def update_target_limit(source_id: str, target_limit: int):
    """Update target limit for a source"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET target_limit = ? WHERE source_id = ?", (target_limit, source_id))
        await db.commit()
        return {"message": f"Target limit updated to {target_limit}"}

@api_router.patch("/sources/{source_id}/keep-non-vessel")
async def update_keep_non_vessel(source_id: str, keep_non_vessel: bool):
    """Update keep non-vessel targets setting"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET keep_non_vessel_targets = ? WHERE source_id = ?", (keep_non_vessel, source_id))
        await db.commit()
        return {"message": f"Keep non-vessel targets: {keep_non_vessel}"}

@api_router.delete("/sources/{source_id}")
async def delete_source(source_id: str, delete_data: bool = False):
    """Delete a source and optionally its data"""
    async with aiosqlite.connect(DB_PATH) as db:
        if delete_data:
            # Delete all related data
            await db.execute("DELETE FROM messages WHERE source_id = ?", (source_id,))
            await db.execute("DELETE FROM positions WHERE source_id = ?", (source_id,))
            await db.execute("DELETE FROM text_messages WHERE source_id = ?", (source_id,))
        
        await db.execute("DELETE FROM sources WHERE source_id = ?", (source_id,))
        await db.commit()
    
    # Stop stream if active
    if source_id in active_streams:
        active_streams[source_id]['running'] = False
        del active_streams[source_id]
    
    return {"message": f"Source {source_id} deleted"}

@api_router.post("/sources/disable-all")
async def disable_all_sources():
    """Disable all sources"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET status = 'inactive'")
        await db.commit()
    
    # Stop all active streams
    for stream_id in list(active_streams.keys()):
        active_streams[stream_id]['running'] = False
        del active_streams[stream_id]
    
    return {"message": "All sources disabled"}

@api_router.get("/vessels/active")
async def get_active_vessels(
    limit: int = 1000,
    min_positions: int = 1,
    hours_back: int = 24,
    include_non_vessels: bool = True
):
    """Get active vessels with recent positions"""
    cutoff_time = datetime.now(timezone.utc) - timedelta(hours=hours_back)
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        query = """
            SELECT v.*, e.profile_data, e.image_url, e.enriched_at
            FROM vessels v
            LEFT JOIN vessel_enrichment e ON v.mmsi = e.mmsi
            WHERE v.position_count >= ? AND v.last_seen >= ?
        """
        params = [min_positions, cutoff_time.isoformat()]
        
        if not include_non_vessels:
            query += " AND (v.ship_type IS NULL OR v.ship_type NOT IN (0, 4, 11, 21))"
        
        query += " ORDER BY v.last_seen DESC LIMIT ?"
        params.append(limit)
        
        async with db.execute(query, params) as cursor:
            vessels = []
            async for row in cursor:
                vessel = dict(row)
                # Parse MarineISA profile
                if vessel.get('profile_data'):
                    try:
                        vessel['marinesia_profile'] = json.loads(vessel['profile_data'])
                        del vessel['profile_data']
                    except:
                        pass
                vessels.append(vessel)
            
            return {
                "vessels": vessels,
                "count": len(vessels),
                "filters": {
                    "limit": limit,
                    "min_positions": min_positions,
                    "hours_back": hours_back,
                    "include_non_vessels": include_non_vessels
                }
            }

@api_router.post("/database/clear")
async def clear_database():
    """Clear all AIS data from database"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM positions")
        await db.execute("DELETE FROM messages")
        await db.execute("DELETE FROM text_messages")
        await db.execute("DELETE FROM vessels")
        await db.execute("DELETE FROM vessel_enrichment")
        await db.execute("UPDATE sources SET message_count = 0, target_count = 0, fragment_count = 0")
        await db.commit()
    
    return {"message": "Database cleared successfully"}

@api_router.post("/database/update-spoof-limits")
async def update_spoof_limits():
    """Update spoof limits for all sources"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE sources SET spoof_limit_km = 500.0 WHERE spoof_limit_km IS NULL")
        await db.commit()
    
    return {"message": "Spoof limits updated"}

@api_router.get("/status")
async def get_status():
    """Get system status"""
    async with aiosqlite.connect(DB_PATH) as db:
        # Count vessels
        async with db.execute("SELECT COUNT(*) FROM vessels") as cursor:
            vessel_count = (await cursor.fetchone())[0]
        
        # Count positions
        async with db.execute("SELECT COUNT(*) FROM positions") as cursor:
            position_count = (await cursor.fetchone())[0]
        
        # Count messages
        async with db.execute("SELECT COUNT(*) FROM messages") as cursor:
            message_count = (await cursor.fetchone())[0]
        
        # Count sources
        async with db.execute("SELECT COUNT(*) FROM sources") as cursor:
            source_count = (await cursor.fetchone())[0]
        
        # Count enriched vessels
        async with db.execute("SELECT COUNT(*) FROM vessel_enrichment") as cursor:
            enriched_count = (await cursor.fetchone())[0]
        
        # Get recent activity
        cutoff_time = datetime.now(timezone.utc) - timedelta(hours=1)
        async with db.execute("""
            SELECT COUNT(*) FROM vessels WHERE last_seen >= ?
        """, (cutoff_time.isoformat(),)) as cursor:
            recent_vessels = (await cursor.fetchone())[0]
    
    return {
        "database": {
            "path": str(DB_PATH),
            "vessels": vessel_count,
            "positions": position_count,
            "messages": message_count,
            "sources": source_count,
            "enriched_vessels": enriched_count,
            "recent_vessels_1h": recent_vessels
        },
        "streams": {
            "active": len(active_streams),
            "stream_ids": list(active_streams.keys())
        },
        "marinesia": {
            "enabled": MARINESIA_ENABLED,
            "queue_size": enrichment_queue.qsize()
        },
        "websocket": {
            "connections": len(manager.active_connections)
        }
    }

@api_router.get("/messages/text")
async def get_text_messages(
    limit: int = 100,
    mmsi: Optional[str] = None,
    message_type: Optional[int] = None
):
    """Get text messages"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        query = "SELECT * FROM text_messages WHERE 1=1"
        params = []
        
        if mmsi:
            query += " AND mmsi = ?"
            params.append(mmsi)
        
        if message_type:
            query += " AND message_type = ?"
            params.append(message_type)
        
        query += " ORDER BY timestamp DESC LIMIT ?"
        params.append(limit)
        
        async with db.execute(query, params) as cursor:
            messages = [dict(row) for row in await cursor.fetchall()]
            return {"messages": messages, "count": len(messages)}

@api_router.get("/messages/text/export")
async def export_text_messages():
    """Export text messages to Excel"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT * FROM text_messages ORDER BY timestamp DESC
        """) as cursor:
            messages = [dict(row) for row in await cursor.fetchall()]
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Text Messages"
    
    # Headers
    headers = ["MMSI", "Timestamp", "Message Type", "Text", "Destination MMSI", "Source ID"]
    ws.append(headers)
    
    # Data
    for msg in messages:
        ws.append([
            msg['mmsi'],
            msg['timestamp'],
            msg['message_type'],
            msg['text'],
            msg.get('dest_mmsi', ''),
            msg.get('source_id', '')
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=text_messages.xlsx"}
    )

@api_router.get("/export/xlsx")
async def export_to_excel(
    hours_back: int = 24,
    include_positions: bool = True,
    include_messages: bool = False
):
    """Export data to Excel file"""
    cutoff_time = datetime.now(timezone.utc) - timedelta(hours=hours_back)
    
    wb = Workbook()
    
    # Vessels sheet
    ws_vessels = wb.active
    ws_vessels.title = "Vessels"
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        # Get vessels
        async with db.execute("""
            SELECT * FROM vessels WHERE last_seen >= ? ORDER BY last_seen DESC
        """, (cutoff_time.isoformat(),)) as cursor:
            vessels = [dict(row) for row in await cursor.fetchall()]
        
        if vessels:
            # Headers
            headers = list(vessels[0].keys())
            ws_vessels.append(headers)
            
            # Data
            for vessel in vessels:
                ws_vessels.append(list(vessel.values()))
        
        # Positions sheet
        if include_positions:
            ws_positions = wb.create_sheet("Positions")
            async with db.execute("""
                SELECT * FROM positions WHERE timestamp >= ? ORDER BY timestamp DESC LIMIT 10000
            """, (cutoff_time.isoformat(),)) as cursor:
                positions = [dict(row) for row in await cursor.fetchall()]
            
            if positions:
                headers = list(positions[0].keys())
                ws_positions.append(headers)
                
                for position in positions:
                    ws_positions.append(list(position.values()))
        
        # Messages sheet
        if include_messages:
            ws_messages = wb.create_sheet("Messages")
            async with db.execute("""
                SELECT * FROM messages WHERE timestamp >= ? ORDER BY timestamp DESC LIMIT 5000
            """, (cutoff_time.isoformat(),)) as cursor:
                messages = [dict(row) for row in await cursor.fetchall()]
            
            if messages:
                headers = list(messages[0].keys())
                ws_messages.append(headers)
                
                for message in messages:
                    ws_messages.append(list(message.values()))
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ais_export_{hours_back}h.xlsx"}
    )

@api_router.get("/serial/ports")
async def get_serial_ports():
    """Get available serial ports"""
    ports = serial.tools.list_ports.comports()
    return {"ports": [{"device": port.device, "description": port.description} for port in ports]}

@api_router.get("/vdo/positions")
async def get_vdo_positions(limit: int = 100):
    """Get VDO (own vessel) positions"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT p.*, v.name, v.ship_type_text 
            FROM positions p
            LEFT JOIN vessels v ON p.mmsi = v.mmsi
            WHERE p.source LIKE '%VDO%' OR p.raw LIKE '!AIVDO%'
            ORDER BY p.timestamp DESC LIMIT ?
        """, (limit,)) as cursor:
            positions = [dict(row) for row in await cursor.fetchall()]
            return {"positions": positions, "count": len(positions)}

@api_router.get("/vessels")
async def get_vessels():
    """Get all vessels with enrichment data"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT v.*, e.profile_data, e.image_url, e.enriched_at
            FROM vessels v
            LEFT JOIN vessel_enrichment e ON v.mmsi = e.mmsi
            ORDER BY v.last_seen DESC
        """) as cursor:
            vessels = []
            async for row in cursor:
                vessel = dict(row)
                # Parse MarineISA profile
                if vessel.get('profile_data'):
                    try:
                        vessel['marinesia_profile'] = json.loads(vessel['profile_data'])
                        del vessel['profile_data']
                    except:
                        pass
                vessels.append(vessel)
            
            return {"vessels": vessels, "count": len(vessels)}

@api_router.get("/vessel/{mmsi}")
async def get_vessel(mmsi: str):
    """Get specific vessel with enrichment"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT v.*, e.profile_data, e.image_url, e.enriched_at
            FROM vessels v
            LEFT JOIN vessel_enrichment e ON v.mmsi = e.mmsi
            WHERE v.mmsi = ?
        """, (mmsi,)) as cursor:
            row = await cursor.fetchone()
            if row:
                vessel = dict(row)
                # Parse MarineISA profile
                if vessel.get('profile_data'):
                    try:
                        vessel['marinesia_profile'] = json.loads(vessel['profile_data'])
                        del vessel['profile_data']
                    except:
                        pass
                return vessel
            raise HTTPException(status_code=404, detail="Vessel not found")

@api_router.get("/vessel/{mmsi}/marinesia-status")
async def get_marinesia_status(mmsi: str):
    """Get MarineISA enrichment status for a vessel"""
    if not marinesia_client:
        return {"status": "disabled", "data": None}
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        # Check if vessel has enrichment data
        async with db.execute("""
            SELECT profile_data, image_url, enriched_at 
            FROM vessel_enrichment 
            WHERE mmsi = ?
        """, (mmsi,)) as cursor:
            row = await cursor.fetchone()
            
            if row and row['profile_data']:
                try:
                    profile_data = json.loads(row['profile_data'])
                    # Check if it's a "not found" marker
                    if profile_data.get('not_found'):
                        return {
                            "status": "not_found",
                            "data": None,
                            "checked_at": row['enriched_at']
                        }
                    else:
                        # Has real data
                        return {
                            "status": "found",
                            "data": profile_data.get('data'),
                            "image_url": row['image_url'],
                            "enriched_at": row['enriched_at']
                        }
                except:
                    pass
        
        # Check if in queue
        # For now, if not enriched, assume it's queued
        return {"status": "queued", "data": None}

@api_router.post("/vessel/{mmsi}/enrich-priority")
async def enrich_vessel_priority(mmsi: str):
    """Trigger priority enrichment for a vessel"""
    if not marinesia_client:
        raise HTTPException(status_code=503, detail="MarineISA integration not enabled")
    
    # Add to enrichment queue (with priority by clearing old attempts)
    async with aiosqlite.connect(DB_PATH) as db:
        # Delete old enrichment attempt
        await db.execute("DELETE FROM vessel_enrichment WHERE mmsi = ?", (mmsi,))
        await db.commit()
    
    # Queue for enrichment
    await enrichment_queue.put(mmsi)
    
    return {
        "message": f"Vessel {mmsi} queued for priority enrichment",
        "queue_position": enrichment_queue.qsize()
    }



@api_router.post("/search")
async def search_vessels(query: SearchQuery):
    """Search vessels based on criteria"""
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        sql_query = "SELECT * FROM vessels WHERE 1=1"
        params = []
        
        if query.mmsi:
            sql_query += " AND mmsi LIKE ?"
            params.append(f"%{query.mmsi}%")
        
        if query.vessel_name:
            sql_query += " AND name LIKE ?"
            params.append(f"%{query.vessel_name}%")
        
        if query.ship_type:
            sql_query += " AND ship_type = ?"
            params.append(query.ship_type)
        
        if query.start_time:
            sql_query += " AND last_seen >= ?"
            params.append(query.start_time.isoformat())
        
        if query.end_time:
            sql_query += " AND last_seen <= ?"
            params.append(query.end_time.isoformat())
        
        sql_query += " ORDER BY last_seen DESC LIMIT 1000"
        
        async with db.execute(sql_query, params) as cursor:
            vessels = [dict(row) for row in await cursor.fetchall()]
            return {"vessels": vessels, "count": len(vessels)}

@api_router.get("/positions/recent")
async def get_recent_positions(
    limit: int = 1000,
    hours_back: int = 1
):
    """Get recent positions"""
    cutoff_time = datetime.now(timezone.utc) - timedelta(hours=hours_back)
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT p.*, v.name, v.ship_type_text
            FROM positions p
            LEFT JOIN vessels v ON p.mmsi = v.mmsi
            WHERE p.timestamp >= ?
            ORDER BY p.timestamp DESC LIMIT ?
        """, (cutoff_time.isoformat(), limit)) as cursor:
            positions = [dict(row) for row in await cursor.fetchall()]
            return {"positions": positions, "count": len(positions)}

@api_router.get("/track/{mmsi}")
async def get_vessel_track(
    mmsi: str,
    hours_back: int = 24,
    limit: int = 1000
):
    """Get vessel track (positions over time)"""
    cutoff_time = datetime.now(timezone.utc) - timedelta(hours=hours_back)
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("""
            SELECT * FROM positions 
            WHERE mmsi = ? AND timestamp >= ?
            ORDER BY timestamp ASC LIMIT ?
        """, (mmsi, cutoff_time.isoformat(), limit)) as cursor:
            positions = [dict(row) for row in await cursor.fetchall()]
            return {
                "mmsi": mmsi,
                "track": positions,
                "count": len(positions),
                "hours_back": hours_back
            }

@api_router.get("/history/{mmsi}")
async def get_vessel_history(
    mmsi: str,
    days_back: int = 7,
    include_messages: bool = False
):
    """Get vessel history"""
    cutoff_time = datetime.now(timezone.utc) - timedelta(days=days_back)
    
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        
        # Get vessel info
        async with db.execute("SELECT * FROM vessels WHERE mmsi = ?", (mmsi,)) as cursor:
            vessel = await cursor.fetchone()
            if not vessel:
                raise HTTPException(status_code=404, detail="Vessel not found")
            vessel = dict(vessel)
        
        # Get positions
        async with db.execute("""
            SELECT * FROM positions 
            WHERE mmsi = ? AND timestamp >= ?
            ORDER BY timestamp DESC LIMIT 10000
        """, (mmsi, cutoff_time.isoformat())) as cursor:
            positions = [dict(row) for row in await cursor.fetchall()]
        
        result = {
            "vessel": vessel,
            "positions": positions,
            "position_count": len(positions),
            "days_back": days_back
        }
        
        # Include messages if requested
        if include_messages:
            async with db.execute("""
                SELECT * FROM messages 
                WHERE mmsi = ? AND timestamp >= ?
                ORDER BY timestamp DESC LIMIT 1000
            """, (mmsi, cutoff_time.isoformat())) as cursor:
                messages = [dict(row) for row in await cursor.fetchall()]
                result["messages"] = messages
                result["message_count"] = len(messages)
        
        return result

@api_router.get("/enrichment-status")
async def enrichment_status():
    """Get enrichment statistics"""
    async with aiosqlite.connect(DB_PATH) as db:
        # Total vessels
        async with db.execute("SELECT COUNT(*) FROM vessels") as cursor:
            total = (await cursor.fetchone())[0]
        
        # Enriched vessels
        async with db.execute("SELECT COUNT(*) FROM vessel_enrichment") as cursor:
            enriched = (await cursor.fetchone())[0]
        
        # Queue size
        queue_size = enrichment_queue.qsize()
        
        return {
            "total_vessels": total,
            "enriched_vessels": enriched,
            "enrichment_queue": queue_size,
            "enrichment_percentage": round((enriched / total * 100) if total > 0 else 0, 1),
            "marinesia_enabled": MARINESIA_ENABLED
        }

@api_router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    """WebSocket for real-time updates"""
    await manager.connect(websocket)
    try:
        while True:
            data = await websocket.receive_text()
            # Echo back for now
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# Include router
app.include_router(api_router)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8002)
